import logging
import os
import traceback
from contextlib import asynccontextmanager
from fastapi import FastAPI, Depends, HTTPException, Query, Header, Request
from sqlalchemy.orm import Session
from typing import List, Optional
from . import models, database
from pydantic import BaseModel, Field, model_validator
from datetime import date, datetime, timedelta
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import func, case, or_, text
import calendar
from fastapi.responses import JSONResponse, StreamingResponse
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from .routers import users, auth, financial_freedom, savings, exchange_rates, banking, tink, stripe_billing, bank_transactions, gamification, admin, budget, reconciliation, partner, ai_chat
from datetime import timezone
from .routers.stripe_billing import TRIAL_DAYS
from .database import engine, Base
from .routers.users import User, UserBase, Settings, SettingsBase  # Import User, UserBase, Settings, and SettingsBase models from users router
import json
import re
import httpx
from .logging_utils import make_conditional_print
from .services.subscription_service import SubscriptionService
from .services.gamification_service import GamificationService
from .services.monthly_totals_service import MonthlyTotalsService
from .services.scheduler_service import (
    initialize_scheduler,
    start_scheduler,
    shutdown_scheduler,
    add_job,
)
from .dependencies import get_current_user as get_authenticated_user

# Rate limiting
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded


def get_user_identifier(request: Request) -> str:
    """
    Get user identifier for rate limiting.
    Uses X-User-ID header if available (authenticated requests),
    otherwise falls back to IP address.
    """
    # Try to get user ID from header (set by NextAuth proxy)
    user_id = request.headers.get("X-User-ID")
    if user_id:
        return f"user:{user_id}"

    # Try to get from Authorization header (JWT token)
    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        try:
            import jwt
            token = auth_header[7:]
            # Decode without verification just to get the subject
            # Full verification happens in the endpoint
            payload = jwt.decode(token, options={"verify_signature": False})
            user_id = payload.get("sub")
            if user_id:
                return f"user:{user_id}"
        except Exception:
            pass

    # Fallback to IP address
    return f"ip:{get_remote_address(request)}"

# Sentry for error tracking
import sentry_sdk
from sentry_sdk.integrations.fastapi import FastApiIntegration
from sentry_sdk.integrations.sqlalchemy import SqlalchemyIntegration

logger = logging.getLogger(__name__)

# Security configuration
from .security import (
    SecurityConfig,
    safe_error_response,
    get_allowed_origins,
    sanitize_log_data,
    mask_email,
)

# Validate security configuration at startup (disabled for tests)
if os.getenv("TESTING") != "true":
    try:
        SecurityConfig.validate_startup()
    except ValueError as e:
        logger.critical(str(e))
        # In production, fail hard. In development, warn but continue.
        if SecurityConfig.IS_PRODUCTION:
            raise

# Initialize Sentry - DSN must be set via environment variable
SENTRY_DSN = os.getenv("SENTRY_DSN", "")
if SENTRY_DSN:
    sentry_sdk.init(
        dsn=SENTRY_DSN,
        integrations=[
            FastApiIntegration(transaction_style="endpoint"),
            SqlalchemyIntegration(),
        ],
        traces_sample_rate=0.1,  # 10% of transactions for performance monitoring
        profiles_sample_rate=0.1,
        environment=os.getenv("ENVIRONMENT", "production"),
    )
else:
    logger.warning("SENTRY_DSN not set. Error tracking is disabled.")

# Initialize rate limiter with per-user identification
# This ensures rate limits are applied per-user rather than per-IP
limiter = Limiter(key_func=get_user_identifier)

# Rate limiter wrapper with async check method
class LimiterWithCheck:
    """Wrapper around slowapi Limiter that adds an async check method."""
    
    def __init__(self, base_limiter: Limiter):
        self._limiter = base_limiter
        self._key_func = base_limiter._key_func
    
    async def check(self, rate_limit: str, request):
        """Rate limit check - currently pass-through. TODO: implement properly."""
        pass
    def __getattr__(self, name):
        return getattr(self._limiter, name)

# Wrap the limiter with check capability
limiter = LimiterWithCheck(limiter)


print = make_conditional_print(__name__)


def validate_user_access(user_id_from_path: str, authenticated_user: models.User) -> None:
    """Validate that the authenticated user matches the user_id in the path.

    Supports partner access: partners can access their household's data
    via the primary user's ID.
    """
    if user_id_from_path == authenticated_user.id:
        return
    if user_id_from_path == authenticated_user.household_id:
        return  # Partner accessing household data
    raise HTTPException(
        status_code=403,
        detail="Access denied: You can only access your own data"
    )


@asynccontextmanager
async def lifespan(app: FastAPI):
    """
    Lifespan context manager for FastAPI application.

    Handles:
    - Scheduler initialization and startup
    - Background job registration
    - Graceful shutdown
    """
    # Startup
    logger.info("Starting application...")

    # Initialize and start scheduler
    try:
        scheduler = initialize_scheduler()
        if scheduler:
            start_scheduler()
            logger.info("Scheduler initialized and started")

            # Add Tink sync job (use async function directly with AsyncIOScheduler)
            from .jobs.tink_sync_job import sync_all_tink_connections

            # Get sync interval from environment (default: 6 hours)
            sync_interval_hours = int(os.getenv("TINK_SYNC_INTERVAL_HOURS", "6"))

            add_job(
                func=sync_all_tink_connections,
                trigger='interval',
                hours=sync_interval_hours,
                id='tink_sync_job',
                name='Tink Connection Sync',
                replace_existing=True,
                max_instances=1,
            )

            logger.info(
                f"Tink sync job scheduled: every {sync_interval_hours} hours"
            )
        else:
            logger.info("Scheduler is disabled")

    except Exception as e:
        logger.error(f"Error initializing scheduler: {str(e)}", exc_info=True)

    yield

    # Shutdown
    logger.info("Shutting down application...")

    try:
        shutdown_scheduler(wait=True)
        logger.info("Scheduler shut down successfully")
    except Exception as e:
        logger.error(f"Error shutting down scheduler: {str(e)}", exc_info=True)


app = FastAPI(
    title="FiredUp API",
    description="Personal finance management API",
    version="1.0.0",
    lifespan=lifespan,
)

# Add rate limiter to app state and exception handler
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# Configure CORS - uses secure helper that excludes localhost in production
allowed_origins = get_allowed_origins()
logger.info(f"CORS allowed origins: {allowed_origins}")

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Content-Type", "Authorization", "X-User-ID", "X-Requested-With"],
)

# Create the database tables
Base.metadata.create_all(bind=engine)


# Health check endpoint
@app.get("/health", tags=["Health"])
async def health_check(db: Session = Depends(database.get_db)):
    """
    Health check endpoint for monitoring.
    Verifies database connectivity.
    """
    try:
        # Test database connection
        db.execute(text("SELECT 1"))
        db_status = "healthy"
    except Exception as e:
        db_status = f"unhealthy: {str(e)}"

    return {
        "status": "ok" if db_status == "healthy" else "degraded",
        "database": db_status,
        "timestamp": datetime.utcnow().isoformat(),
    }

# Include routers
app.include_router(auth.router)
app.include_router(users.router)
# Financial freedom with two prefixes:
# - /financial-freedom for web (goes through Next.js API proxy)
# - /internal-api/financial-freedom for mobile (direct backend access)
app.include_router(financial_freedom.router)
app.include_router(financial_freedom.router, prefix="/internal-api")
app.include_router(savings.router)
app.include_router(exchange_rates.router)
app.include_router(banking.router)
app.include_router(tink.router)
app.include_router(bank_transactions.router)
app.include_router(stripe_billing.router)
# Mobile billing access (direct backend, bypasses Next.js routing)
app.include_router(stripe_billing.router, prefix="/internal-api")
# Gamification system
app.include_router(gamification.router)
app.include_router(gamification.router, prefix="/internal-api")
# Admin endpoints (audit logs, etc.)
app.include_router(admin.router)
# Budget planning (annual budgets from onboarding)
app.include_router(budget.router)
# Reconciliation (bank/manual transaction deduplication)
app.include_router(reconciliation.router)
# Partner access (invite, accept, status, unlink)
app.include_router(partner.router)
# AI Chat (WebSocket + REST)
app.include_router(ai_chat.router)

# Loan models
VALID_LOAN_TYPES = ["mortgage", "car", "personal", "student", "credit_card", "cash_loan", "installment", "leasing", "overdraft", "other"]

class LoanBase(BaseModel):
    loan_type: str = Field(..., description="Type of loan")
    description: str = Field(..., min_length=1, max_length=100, description="Loan description")
    principal_amount: float = Field(..., gt=0, description="Original loan amount")
    remaining_balance: float = Field(..., ge=0, description="Current remaining balance")
    interest_rate: float = Field(..., ge=0, le=100, description="Annual interest rate as percentage")
    monthly_payment: float = Field(..., ge=0, description="Monthly payment amount")
    start_date: date = Field(..., description="Loan start date")
    term_months: int = Field(..., gt=0, description="Loan term in months")
    due_day: int | None = Field(default=1, ge=1, le=31, description="Day of month when payment is due (1-31)")
    # Polish prepayment regulations (since 2022, banks cannot charge fees for first 3 years)
    overpayment_fee_percent: float | None = Field(default=0, ge=0, le=10, description="Prepayment fee percentage (0-10%)")
    overpayment_fee_waived_until: date | None = Field(default=None, description="Date until prepayment fees are waived")

    @model_validator(mode='after')
    def validate_loan(self):
        # Validate loan type
        if self.loan_type not in VALID_LOAN_TYPES:
            raise ValueError(f"Invalid loan type. Must be one of: {', '.join(VALID_LOAN_TYPES)}")

        # Validate remaining balance doesn't exceed principal
        if self.remaining_balance > self.principal_amount:
            raise ValueError("Remaining balance cannot exceed principal amount")

        # Validate monthly payment isn't greater than principal
        if self.monthly_payment > self.principal_amount:
            raise ValueError("Monthly payment cannot exceed principal amount")

        return self

class LoanCreate(LoanBase):
    pass

class Loan(LoanBase):
    id: int
    user_id: str
    created_at: datetime
    updated_at: datetime | None = None

    class Config:
        from_attributes = True
        json_encoders = {
            datetime: lambda v: v.isoformat()
        }


# Loan Payment models
VALID_PAYMENT_TYPES = ["regular", "overpayment"]

class LoanPaymentBase(BaseModel):
    amount: float = Field(..., gt=0, description="Payment amount")
    payment_date: date = Field(..., description="Date of payment")
    payment_type: str = Field(..., description="Type: 'regular' or 'overpayment'")
    covers_month: int | None = Field(default=None, ge=1, le=12, description="Month this payment covers (1-12)")
    covers_year: int | None = Field(default=None, ge=2000, le=2100, description="Year this payment covers")
    notes: str | None = Field(default=None, max_length=500, description="Optional notes")

    @model_validator(mode='after')
    def validate_payment(self):
        if self.payment_type not in VALID_PAYMENT_TYPES:
            raise ValueError(f"Invalid payment type. Must be one of: {', '.join(VALID_PAYMENT_TYPES)}")
        # Regular payments should have covers_month and covers_year
        if self.payment_type == "regular" and (self.covers_month is None or self.covers_year is None):
            raise ValueError("Regular payments must specify covers_month and covers_year")
        return self

class LoanPaymentCreate(LoanPaymentBase):
    pass

class LoanPayment(LoanPaymentBase):
    id: int
    loan_id: int
    user_id: str
    created_at: datetime
    updated_at: datetime | None = None

    class Config:
        from_attributes = True
        json_encoders = {
            datetime: lambda v: v.isoformat()
        }


class LoanPaymentResponse(BaseModel):
    """Extended response for loan payment that includes celebration data."""
    payment: LoanPayment
    new_balance: float
    loan_paid_off: bool = False
    celebration: dict | None = None
    xp_earned: int = 0


# Expense models
class ExpenseBase(BaseModel):
    category: str
    description: str
    amount: float
    is_recurring: bool = False
    date: date  # Start date (for recurring) or occurrence date (for one-off)
    end_date: date | None = None  # Optional end date for recurring items (null = forever)
    owner: str | None = Field(default=None, description="Expense owner: 'self', 'partner' (null = 'self')")
    source: str = "manual"  # "manual" or "bank_import"
    bank_transaction_id: int | None = None  # Link to bank transaction if created from bank import
    reconciliation_status: str = "unreviewed"  # For future reconciliation features

class ExpenseCreate(ExpenseBase):
    pass

class Expense(ExpenseBase):
    id: int
    user_id: str
    created_at: datetime
    updated_at: datetime | None = None

    class Config:
        from_attributes = True
        json_encoders = {
            datetime: lambda v: v.isoformat()
        }

# Income models
# Employment types for Polish tax calculation
EMPLOYMENT_TYPES = ["uop", "b2b", "zlecenie", "dzielo", "other"]

class IncomeBase(BaseModel):
    category: str
    description: str
    amount: float  # Net amount (netto) - what you receive after tax
    is_recurring: bool = False
    date: date  # Start date (for recurring) or occurrence date (for one-off)
    end_date: date | None = None  # Optional end date for recurring items (null = forever)
    # Polish employment type for tax calculation
    employment_type: str | None = Field(default=None, description="Employment type: uop (umowa o pracę), b2b, zlecenie, dzielo, other")
    gross_amount: float | None = Field(default=None, ge=0, description="Gross amount (brutto) before tax")
    is_gross: bool = Field(default=False, description="Whether the entered amount was gross (true) or net (false)")
    kup_type: str | None = Field(default=None, description="KUP type: 'standard', 'author_50', 'none' (null = use global setting)")
    owner: str | None = Field(default=None, description="Income owner: 'self', 'partner' (null = 'self')")
    source: str = "manual"  # "manual" or "bank_import"
    bank_transaction_id: int | None = None  # Link to bank transaction if created from bank import
    reconciliation_status: str = "unreviewed"  # For future reconciliation features

class IncomeCreate(IncomeBase):
    pass

class Income(IncomeBase):
    id: int
    user_id: str
    created_at: datetime
    updated_at: datetime | None = None

    class Config:
        from_attributes = True
        json_encoders = {
            datetime: lambda v: v.isoformat()
        }

# Activity models
class ActivityBase(BaseModel):
    entity_type: str
    operation_type: str
    entity_id: int
    previous_values: Optional[dict] = None
    new_values: Optional[dict] = None

class ActivityCreate(ActivityBase):
    pass

class Activity(ActivityBase):
    id: int
    user_id: str
    timestamp: datetime


class ImportPayload(BaseModel):
    data: dict
    clear_existing: bool = False

# Monthly Budget Report models
class MonthlyBudgetData(BaseModel):
    incomes: List[Income]
    expenses: List[Expense]
    loan_payments: List[Loan]
    totals: dict

class YearlyBudgetReport(BaseModel):
    months: dict[str, MonthlyBudgetData]

# Insights models
class InsightMetric(BaseModel):
    label: str
    value: str
    trend: str  # "up" | "down" | "stable"

class Insight(BaseModel):
    type: str  # "observation" | "recommendation" | "alert" | "achievement"
    title: str
    description: str
    priority: str  # "high" | "medium" | "low"
    actionItems: list[str]
    metrics: list[InsightMetric] = []

class CategoryInsights(BaseModel):
    insights: list[Insight]
    status: str  # "good" | "can_be_improved" | "ok" | "bad"

class InsightsResponse(BaseModel):
    categories: dict[str, list[Insight]]
    status: dict[str, str]
    metadata: dict

class InsightsCache(BaseModel):
    id: int
    userId: str
    insights: InsightsResponse
    financialSnapshot: dict
    createdAt: datetime
    isStale: bool

    class Config:
        from_attributes = True
        json_encoders = {
            datetime: lambda v: v.isoformat()
        }

# Define SettingsCreate model
class SettingsCreate(SettingsBase):
    pass

# User endpoints
# DEPRECATED: This endpoint is kept for backward compatibility only.
# Use /users/me from the users router which properly validates X-Internal-Secret.
# This endpoint will be removed in a future release.
@app.get("/users/me", response_model=User, include_in_schema=False)
def get_or_create_current_user_deprecated(
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """
    DEPRECATED: Get or create current user.
    This endpoint now requires proper authentication via X-Internal-Secret header
    or Bearer token. Use /users/me from the users router instead.
    """
    # User already authenticated and fetched by dependency
    # Check if subscription exists, create trial if not
    subscription = db.query(models.Subscription).filter(
        models.Subscription.user_id == current_user.household_id
    ).first()

    if not subscription:
        now = datetime.now(timezone.utc)
        trial_end = now + timedelta(days=TRIAL_DAYS)
        subscription = models.Subscription(
            user_id=current_user.household_id,
            status="trialing",
            plan_type="trial",
            trial_start=now,
            trial_end=trial_end,
            is_lifetime=False,
            cancel_at_period_end=False,
        )
        db.add(subscription)
        db.commit()
        print(f"[FastAPI] Created trial subscription for user (ends: {trial_end})")

    return current_user

@app.post("/users", response_model=User)
def create_user(user: UserBase, db: Session = Depends(database.get_db)):
    print(f"[FastAPI] Creating user with email: {user.email}")
    
    # Check if user already exists by email
    existing_user = db.query(models.User).filter(
        (models.User.email == user.email) | (models.User.id == user.email)
    ).first()
    
    if existing_user:
        print(f"[FastAPI] User already exists with email: {user.email}")
        return existing_user
    
    # Create new user with email as ID
    db_user = models.User(
        id=user.email,  # Using email as ID
        email=user.email,
        name=user.name
    )
    
    try:
        db.add(db_user)
        db.commit()
        db.refresh(db_user)
        
        # Create default settings for the new user
        default_settings = models.Settings(
            user_id=db_user.id,
            language="en",
            currency="USD",
            ai={"apiKey": None}
        )
        db.add(default_settings)
        db.commit()
        
        print(f"[FastAPI] Created new user and settings: {db_user}")
        return db_user
    except Exception as e:
        db.rollback()
        print(f"[FastAPI] Error creating user: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# Settings endpoints
@app.get("/users/{user_id}/settings", response_model=Settings)
def get_user_settings(
    user_id: str,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Get settings for the authenticated user."""
    validate_user_access(user_id, current_user)

    try:
        print(f"[FastAPI] Getting settings for user: {current_user.household_id}")

        # Get or create settings
        settings = db.query(models.Settings).filter(models.Settings.user_id == current_user.household_id).first()

        if not settings:
            print(f"[FastAPI] No settings found for user: {current_user.household_id}, creating default settings")
            # Create default settings
            settings = models.Settings(
                user_id=current_user.household_id,
                language="en",
                currency="USD",
                ai={"apiKey": None}
            )
            db.add(settings)
            try:
                db.commit()
                db.refresh(settings)
                print(f"[FastAPI] Created default settings: {settings}")
            except Exception as e:
                db.rollback()
                print(f"[FastAPI] Error creating settings: {str(e)}")
                raise HTTPException(status_code=500, detail=f"Error creating settings: {str(e)}")

        print(f"[FastAPI] Returning settings: {settings}")
        return settings
    except HTTPException:
        raise
    except Exception as e:
        print(f"[FastAPI] Error in get_user_settings: {str(e)}")
        import traceback
        print(f"[FastAPI] Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/users/{user_id}/settings", response_model=Settings)
def update_user_settings(
    user_id: str,
    settings: SettingsCreate,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Update settings for the authenticated user."""
    validate_user_access(user_id, current_user)

    try:
        print(f"[FastAPI] Updating settings for user: {current_user.household_id}")
        db_settings = db.query(models.Settings).filter(models.Settings.user_id == current_user.household_id).first()

        if not db_settings:
            print(f"[FastAPI] No settings found for user: {current_user.household_id}, creating new settings")
            db_settings = models.Settings(user_id=current_user.household_id)
            db.add(db_settings)

        for key, value in settings.model_dump().items():
            setattr(db_settings, key, value)

        db.commit()
        db.refresh(db_settings)
        print(f"[FastAPI] Updated settings: {db_settings}")
        return db_settings
    except HTTPException:
        raise
    except Exception as e:
        print(f"[FastAPI] Error in update_user_settings: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# Loan endpoints
@app.get("/loans", response_model=List[Loan])
def get_loans(
    include_archived: bool = False,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Get all loans for the authenticated user. Excludes archived loans by default."""
    query = db.query(models.Loan).filter(models.Loan.user_id == current_user.household_id)
    if not include_archived:
        query = query.filter(
            (models.Loan.is_archived == False) | (models.Loan.is_archived == None)
        )
    loans = query.all()
    return loans


@app.post("/loans/{loan_id}/archive")
def archive_loan(
    loan_id: int,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Archive a paid-off loan."""
    loan = db.query(models.Loan).filter(
        models.Loan.id == loan_id,
        models.Loan.user_id == current_user.household_id
    ).first()
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")

    # Only allow archiving paid-off loans
    if loan.remaining_balance > 0:
        raise HTTPException(status_code=400, detail="Only paid-off loans can be archived")

    loan.is_archived = True
    db.commit()
    return {"success": True, "message": "Loan archived successfully"}

@app.get("/loans/{loan_id}", response_model=Loan)
def get_loan(
    loan_id: int,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Get a specific loan by ID for the authenticated user."""
    loan = db.query(models.Loan).filter(
        models.Loan.id == loan_id,
        models.Loan.user_id == current_user.household_id
    ).first()
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    return loan

@app.post("/loans", response_model=Loan)
def create_loan(
    loan: LoanCreate,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Create a new loan for the authenticated user."""
    # Check subscription limits
    can_add, message = SubscriptionService.can_add_loan(current_user.household_id, db)
    if not can_add:
        raise HTTPException(status_code=403, detail=message)

    db_loan = models.Loan(
        user_id=current_user.household_id,
        loan_type=loan.loan_type,
        description=loan.description,
        principal_amount=loan.principal_amount,
        remaining_balance=loan.remaining_balance,
        interest_rate=loan.interest_rate,
        monthly_payment=loan.monthly_payment,
        start_date=loan.start_date,
        term_months=loan.term_months,
        due_day=loan.due_day
    )
    db.add(db_loan)
    db.commit()
    db.refresh(db_loan)
    return db_loan

@app.put("/loans/{loan_id}", response_model=Loan)
def update_loan(
    loan_id: int,
    loan: LoanCreate,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Update a loan owned by the authenticated user."""
    db_loan = db.query(models.Loan).filter(
        models.Loan.id == loan_id,
        models.Loan.user_id == current_user.household_id
    ).first()

    if not db_loan:
        raise HTTPException(status_code=404, detail="Loan not found")

    for key, value in loan.model_dump().items():
        setattr(db_loan, key, value)

    db.commit()
    db.refresh(db_loan)
    return db_loan

@app.delete("/users/{user_id}/loans/{loan_id}")
def delete_loan(
    user_id: str,
    loan_id: int,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Delete a loan owned by the authenticated user."""
    validate_user_access(user_id, current_user)

    db_loan = db.query(models.Loan).filter(
        models.Loan.id == loan_id,
        models.Loan.user_id == current_user.household_id
    ).first()

    if not db_loan:
        raise HTTPException(status_code=404, detail="Loan not found")

    db.delete(db_loan)
    db.commit()
    return {"message": "Loan deleted successfully"}


# Loan Payment endpoints
@app.get("/users/{user_id}/loans/{loan_id}/payments", response_model=List[LoanPayment])
def get_loan_payments(
    user_id: str,
    loan_id: int,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Get all payments for a specific loan."""
    validate_user_access(user_id, current_user)

    # Verify loan exists and belongs to user
    db_loan = db.query(models.Loan).filter(
        models.Loan.id == loan_id,
        models.Loan.user_id == current_user.household_id
    ).first()
    if not db_loan:
        raise HTTPException(status_code=404, detail="Loan not found")

    payments = db.query(models.LoanPayment).filter(
        models.LoanPayment.loan_id == loan_id
    ).order_by(models.LoanPayment.payment_date.desc()).all()
    return payments


@app.post("/users/{user_id}/loans/{loan_id}/payments", response_model=LoanPaymentResponse)
def create_loan_payment(
    user_id: str,
    loan_id: int,
    payment: LoanPaymentCreate,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Create a payment for a loan and update remaining balance."""
    validate_user_access(user_id, current_user)

    # Verify loan exists and belongs to user
    db_loan = db.query(models.Loan).filter(
        models.Loan.id == loan_id,
        models.Loan.user_id == current_user.household_id
    ).first()
    if not db_loan:
        raise HTTPException(status_code=404, detail="Loan not found")

    # For regular payments, check if payment already exists for this month
    if payment.payment_type == "regular":
        existing = db.query(models.LoanPayment).filter(
            models.LoanPayment.loan_id == loan_id,
            models.LoanPayment.payment_type == "regular",
            models.LoanPayment.covers_month == payment.covers_month,
            models.LoanPayment.covers_year == payment.covers_year
        ).first()
        if existing:
            raise HTTPException(
                status_code=400,
                detail=f"Payment for {payment.covers_month}/{payment.covers_year} already exists"
            )

    # Create payment record
    db_payment = models.LoanPayment(
        **payment.model_dump(),
        loan_id=loan_id,
        user_id=current_user.household_id
    )
    db.add(db_payment)

    # Update loan remaining balance
    db_loan.remaining_balance = max(0, db_loan.remaining_balance - payment.amount)

    db.commit()
    db.refresh(db_payment)
    db.refresh(db_loan)

    # Trigger gamification - award XP for loan payment
    xp_earned = 0
    celebration = None
    try:
        is_overpayment = payment.payment_type == "overpayment"
        xp_earned, _, celebration = GamificationService.on_loan_payment(
            current_user.id, db_payment.id, loan_id, is_overpayment, db
        )
    except Exception as gam_error:
        print(f"[FastAPI] Gamification error (non-blocking): {gam_error}")

    # Return extended response with celebration data
    return LoanPaymentResponse(
        payment=LoanPayment.model_validate(db_payment),
        new_balance=db_loan.remaining_balance,
        loan_paid_off=db_loan.remaining_balance <= 0,
        celebration=celebration,
        xp_earned=xp_earned,
    )


@app.delete("/users/{user_id}/loans/{loan_id}/payments/{payment_id}")
def delete_loan_payment(
    user_id: str,
    loan_id: int,
    payment_id: int,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Delete (undo) a loan payment and restore the balance."""
    validate_user_access(user_id, current_user)

    # Verify loan exists and belongs to user
    db_loan = db.query(models.Loan).filter(
        models.Loan.id == loan_id,
        models.Loan.user_id == current_user.household_id
    ).first()
    if not db_loan:
        raise HTTPException(status_code=404, detail="Loan not found")

    # Find the payment
    db_payment = db.query(models.LoanPayment).filter(
        models.LoanPayment.id == payment_id,
        models.LoanPayment.loan_id == loan_id
    ).first()
    if not db_payment:
        raise HTTPException(status_code=404, detail="Payment not found")

    # Restore the balance (capped at principal_amount to avoid validation errors)
    db_loan.remaining_balance = min(
        db_loan.remaining_balance + db_payment.amount,
        db_loan.principal_amount
    )

    # Delete the payment
    db.delete(db_payment)
    db.commit()

    return {"message": "Payment deleted and balance restored"}


# Mobile API - Loan endpoints with /internal-api prefix (bypasses Next.js routing)
@app.get("/internal-api/loans", response_model=List[Loan])
def get_loans_internal(
    include_archived: bool = False,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Get all loans for the authenticated user (mobile API)."""
    query = db.query(models.Loan).filter(models.Loan.user_id == current_user.household_id)
    if not include_archived:
        query = query.filter(
            (models.Loan.is_archived == False) | (models.Loan.is_archived == None)
        )
    loans = query.all()
    return loans


@app.post("/internal-api/loans", response_model=Loan)
def create_loan_internal(
    loan: LoanCreate,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Create a new loan for the authenticated user (mobile API)."""
    can_add, message = SubscriptionService.can_add_loan(current_user.household_id, db)
    if not can_add:
        raise HTTPException(status_code=403, detail=message)

    db_loan = models.Loan(
        user_id=current_user.household_id,
        loan_type=loan.loan_type,
        description=loan.description,
        principal_amount=loan.principal_amount,
        remaining_balance=loan.remaining_balance,
        interest_rate=loan.interest_rate,
        monthly_payment=loan.monthly_payment,
        start_date=loan.start_date,
        term_months=loan.term_months,
        due_day=loan.due_day
    )
    db.add(db_loan)
    db.commit()
    db.refresh(db_loan)
    return db_loan


@app.get("/internal-api/loans/{loan_id}", response_model=Loan)
def get_loan_internal(
    loan_id: int,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Get a specific loan by ID (mobile API)."""
    loan = db.query(models.Loan).filter(
        models.Loan.id == loan_id,
        models.Loan.user_id == current_user.household_id
    ).first()
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    return loan


@app.post("/internal-api/loans/{loan_id}/archive")
def archive_loan_internal(
    loan_id: int,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Archive a paid-off loan (mobile API)."""
    loan = db.query(models.Loan).filter(
        models.Loan.id == loan_id,
        models.Loan.user_id == current_user.household_id
    ).first()
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")

    if loan.remaining_balance > 0:
        raise HTTPException(status_code=400, detail="Only paid-off loans can be archived")

    loan.is_archived = True
    db.commit()
    return {"success": True, "message": "Loan archived successfully"}


# Expense endpoints
@app.get("/users/{user_id}/expenses", response_model=List[Expense])
def get_user_expenses(
    user_id: str,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Get all expenses for the authenticated user."""
    validate_user_access(user_id, current_user)

    try:
        print(f"[FastAPI] Getting expenses for user: {current_user.household_id}")
        expenses = db.query(models.Expense).filter(models.Expense.user_id == current_user.household_id).all()
        print(f"[FastAPI] Found {len(expenses)} expenses")
        return expenses
    except HTTPException:
        raise
    except Exception as e:
        print(f"[FastAPI] Error in get_user_expenses: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/users/{user_id}/expenses", response_model=Expense)
def create_expense(
    user_id: str,
    expense: ExpenseCreate,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Create an expense for the authenticated user."""
    validate_user_access(user_id, current_user)

    try:
        print(f"[FastAPI] Creating expense for user: {current_user.household_id}")
        # Check subscription limits
        can_add, message = SubscriptionService.can_add_expense(current_user.household_id, db)
        if not can_add:
            raise HTTPException(status_code=403, detail=message)

        expense_data = expense.model_dump()
        if expense_data.get("owner") is None:
            expense_data["owner"] = "partner" if current_user.is_partner else "self"
        db_expense = models.Expense(**expense_data, user_id=current_user.household_id)
        db.add(db_expense)
        db.commit()
        db.refresh(db_expense)

        # Trigger gamification - award XP for logging expense
        try:
            GamificationService.on_expense_logged(current_user.id, db_expense.id, db)
        except Exception as gam_error:
            print(f"[FastAPI] Gamification error (non-blocking): {gam_error}")

        print(f"[FastAPI] Created expense: {db_expense}")
        return db_expense
    except HTTPException:
        raise
    except Exception as e:
        print(f"[FastAPI] Error in create_expense: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/users/{user_id}/expenses/{expense_id}", response_model=Expense)
def update_expense(
    user_id: str,
    expense_id: int,
    expense: ExpenseCreate,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Update an expense for the authenticated user."""
    validate_user_access(user_id, current_user)

    try:
        print(f"[FastAPI] Updating expense {expense_id} for user: {current_user.household_id}")
        db_expense = db.query(models.Expense).filter(
            models.Expense.id == expense_id,
            models.Expense.user_id == current_user.household_id
        ).first()

        if not db_expense:
            print(f"[FastAPI] Expense not found with ID: {expense_id}")
            raise HTTPException(status_code=404, detail="Expense not found")

        for key, value in expense.model_dump().items():
            setattr(db_expense, key, value)

        db.commit()
        db.refresh(db_expense)
        print(f"[FastAPI] Updated expense: {db_expense}")
        return db_expense
    except HTTPException:
        raise
    except Exception as e:
        print(f"[FastAPI] Error in update_expense: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/users/{user_id}/expenses/{expense_id}")
def delete_expense(
    user_id: str,
    expense_id: int,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Delete an expense for the authenticated user."""
    validate_user_access(user_id, current_user)

    try:
        print(f"[FastAPI] Deleting expense {expense_id} for user: {current_user.household_id}")
        db_expense = db.query(models.Expense).filter(
            models.Expense.id == expense_id,
            models.Expense.user_id == current_user.household_id
        ).first()

        if not db_expense:
            print(f"[FastAPI] Expense not found with ID: {expense_id}")
            raise HTTPException(status_code=404, detail="Expense not found")

        # Clear any bank transaction references before deleting
        # This handles the case where bank_transactions.linked_expense_id points to this expense
        bank_txs = db.query(models.BankTransaction).filter(
            models.BankTransaction.linked_expense_id == expense_id,
            models.BankTransaction.user_id == current_user.household_id
        ).all()

        for bank_tx in bank_txs:
            print(f"[FastAPI] Clearing linked_expense_id from bank transaction {bank_tx.id}")
            bank_tx.linked_expense_id = None

        # Flush changes to clear FK references before deleting
        if bank_txs:
            db.flush()

        db.delete(db_expense)
        db.commit()
        print(f"[FastAPI] Deleted expense: {db_expense}")
        return {"message": "Expense deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        print(f"[FastAPI] Error in delete_expense: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/users/{user_id}/expenses/monthly")
def get_monthly_expenses(
    user_id: str,
    month: Optional[str] = None,  # Format: YYYY-MM (e.g., "2026-02")
    include_bank: bool = True,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """
    Get monthly expenses with bank/manual breakdown.

    Combines bank-backed and manual entries, excludes duplicates.
    If month not specified, uses current month.

    Returns:
        - total: Combined total (bank + manual, after deduplication)
        - from_bank: Total from bank transactions
        - from_manual: Total from manual entries (deduplicated)
        - breakdown: Detailed counts
        - month: YYYY-MM format
    """
    validate_user_access(user_id, current_user)

    try:
        print(f"[FastAPI] Getting monthly expenses for user: {current_user.household_id}, month: {month}")

        # Parse month parameter or use current month
        if month:
            try:
                year, month_num = map(int, month.split('-'))
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
        else:
            today = datetime.now()
            year = today.year
            month_num = today.month
            month = today.strftime("%Y-%m")

        # Use MonthlyTotalsService for deduplication logic
        totals = MonthlyTotalsService.calculate_monthly_expenses(
            user_id=current_user.household_id,
            year=year,
            month=month_num,
            db=db
        )

        print(f"[FastAPI] Monthly expenses: total={totals['total']}, from_bank={totals['from_bank']}, from_manual={totals['from_manual']}")

        return {
            "month": month,
            "total": float(totals["total"]),
            "from_bank": float(totals["from_bank"]),
            "from_manual": float(totals["from_manual"]),
            "breakdown": {
                "bank_count": totals["bank_count"],
                "manual_count": totals["manual_count"],
                "duplicate_count": totals["duplicate_count"],
                "unreviewed_count": totals["unreviewed_count"]
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"[FastAPI] Error in get_monthly_expenses: {str(e)}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# Income endpoints
@app.get("/users/{user_id}/income", response_model=List[Income])
def get_user_income(
    user_id: str,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Get all income entries for the authenticated user."""
    validate_user_access(user_id, current_user)

    try:
        print(f"[FastAPI] Getting income for user: {current_user.household_id}")
        income = db.query(models.Income).filter(models.Income.user_id == current_user.household_id).all()
        print(f"[FastAPI] Found {len(income)} income entries")
        return income
    except HTTPException:
        raise
    except Exception as e:
        print(f"[FastAPI] Error in get_user_income: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/users/{user_id}/income", response_model=Income)
def create_income(
    user_id: str,
    income: IncomeCreate,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Create an income entry for the authenticated user."""
    validate_user_access(user_id, current_user)

    try:
        print(f"[FastAPI] Creating income for user: {current_user.household_id}")
        # Check subscription limits
        can_add, message = SubscriptionService.can_add_income(current_user.household_id, db)
        if not can_add:
            raise HTTPException(status_code=403, detail=message)

        income_data = income.model_dump()
        if income_data.get("owner") is None:
            income_data["owner"] = "partner" if current_user.is_partner else "self"
        db_income = models.Income(**income_data, user_id=current_user.household_id)
        db.add(db_income)
        db.commit()
        db.refresh(db_income)

        # Trigger gamification - award XP for logging income
        try:
            GamificationService.on_income_logged(current_user.id, db_income.id, db)
        except Exception as gam_error:
            print(f"[FastAPI] Gamification error (non-blocking): {gam_error}")

        print(f"[FastAPI] Created income: {db_income}")
        return db_income
    except HTTPException:
        raise
    except Exception as e:
        print(f"[FastAPI] Error in create_income: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/users/{user_id}/income/{income_id}", response_model=Income)
def update_income(
    user_id: str,
    income_id: int,
    income: IncomeCreate,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Update an income entry for the authenticated user."""
    validate_user_access(user_id, current_user)

    try:
        print(f"[FastAPI] Updating income {income_id} for user: {current_user.household_id}")
        db_income = db.query(models.Income).filter(
            models.Income.id == income_id,
            models.Income.user_id == current_user.household_id
        ).first()

        if not db_income:
            print(f"[FastAPI] Income not found with ID: {income_id}")
            raise HTTPException(status_code=404, detail="Income not found")

        for key, value in income.model_dump().items():
            setattr(db_income, key, value)

        db.commit()
        db.refresh(db_income)
        print(f"[FastAPI] Updated income: {db_income}")
        return db_income
    except HTTPException:
        raise
    except Exception as e:
        print(f"[FastAPI] Error in update_income: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/users/{user_id}/income/{income_id}")
def delete_income(
    user_id: str,
    income_id: int,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Delete an income entry for the authenticated user."""
    validate_user_access(user_id, current_user)

    try:
        print(f"[FastAPI] Deleting income {income_id} for user: {current_user.household_id}")
        db_income = db.query(models.Income).filter(
            models.Income.id == income_id,
            models.Income.user_id == current_user.household_id
        ).first()

        if not db_income:
            print(f"[FastAPI] Income not found with ID: {income_id}")
            raise HTTPException(status_code=404, detail="Income not found")

        # Clear any bank transaction references before deleting
        # This handles the case where bank_transactions.linked_income_id points to this income
        bank_txs = db.query(models.BankTransaction).filter(
            models.BankTransaction.linked_income_id == income_id,
            models.BankTransaction.user_id == current_user.household_id
        ).all()

        for bank_tx in bank_txs:
            print(f"[FastAPI] Clearing linked_income_id from bank transaction {bank_tx.id}")
            bank_tx.linked_income_id = None

        # Flush changes to clear FK references before deleting
        if bank_txs:
            db.flush()

        db.delete(db_income)
        db.commit()
        print(f"[FastAPI] Deleted income: {db_income}")
        return {"message": "Income deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        print(f"[FastAPI] Error in delete_income: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/users/{user_id}/income/monthly")
def get_monthly_income(
    user_id: str,
    month: Optional[str] = None,  # Format: YYYY-MM (e.g., "2026-02")
    include_bank: bool = True,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """
    Get monthly income with bank/manual breakdown.

    Combines bank-backed and manual entries, excludes duplicates.
    If month not specified, uses current month.

    Returns:
        - total: Combined total (bank + manual, after deduplication)
        - from_bank: Total from bank transactions
        - from_manual: Total from manual entries (deduplicated)
        - breakdown: Detailed counts
        - month: YYYY-MM format
    """
    validate_user_access(user_id, current_user)

    try:
        print(f"[FastAPI] Getting monthly income for user: {current_user.household_id}, month: {month}")

        # Parse month parameter or use current month
        if month:
            try:
                year, month_num = map(int, month.split('-'))
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
        else:
            today = datetime.now()
            year = today.year
            month_num = today.month
            month = today.strftime("%Y-%m")

        # Use MonthlyTotalsService for deduplication logic
        totals = MonthlyTotalsService.calculate_monthly_income(
            user_id=current_user.household_id,
            year=year,
            month=month_num,
            db=db
        )

        print(f"[FastAPI] Monthly income: total={totals['total']}, from_bank={totals['from_bank']}, from_manual={totals['from_manual']}")

        return {
            "month": month,
            "total": float(totals["total"]),
            "from_bank": float(totals["from_bank"]),
            "from_manual": float(totals["from_manual"]),
            "breakdown": {
                "bank_count": totals["bank_count"],
                "manual_count": totals["manual_count"],
                "duplicate_count": totals["duplicate_count"],
                "unreviewed_count": totals["unreviewed_count"]
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"[FastAPI] Error in get_monthly_income: {str(e)}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# ── Tax calculation (stateless utility, no auth required) ──────────────────────

class TaxCalculationRequest(BaseModel):
    gross_monthly: float = Field(gt=0)
    employment_type: str  # uop, b2b, zlecenie, dzielo, other
    use_authors_costs: bool = False
    ppk_employee_rate: float = 0.0  # 0 if not enrolled
    kup_type: str | None = None  # "standard", "author_50", "none" (null = use global use_authors_costs)

class TaxBreakdown(BaseModel):
    zus: float
    ppk: float
    kup: float
    health: float
    pit: float

class TaxCalculationResponse(BaseModel):
    gross: float
    net: float
    breakdown: TaxBreakdown


def calculate_net_from_gross(
    gross_monthly: float,
    employment_type: str,
    use_authors_costs: bool = False,
    ppk_employee_rate: float = 0.0,
    kup_type: str | None = None,
) -> TaxCalculationResponse:
    """
    Calculate monthly net income from gross for Polish employment types.
    Constants based on 2026 Polish tax law.
    """
    # Tax constants
    ZUS_PENSION = 0.0976
    ZUS_DISABILITY = 0.015
    ZUS_SICKNESS = 0.0245
    ZUS_TOTAL = ZUS_PENSION + ZUS_DISABILITY + ZUS_SICKNESS  # 13.71%
    HEALTH_RATE = 0.09
    KUP_STANDARD_UOP = 250.0  # PLN/month
    KUP_AUTHORS_CAP_MONTHLY = 10000.0  # 120,000/year
    PIT_RATE_1 = 0.12
    PIT_RATE_2 = 0.32
    PIT_THRESHOLD_MONTHLY = 10000.0  # 120,000/year
    TAX_FREE_MONTHLY = 2500.0  # 30,000/year
    B2B_LINEAR_RATE = 0.19

    zus = 0.0
    ppk = 0.0
    kup = 0.0
    health = 0.0
    pit = 0.0

    if employment_type == "other":
        # No deductions — gross equals net
        return TaxCalculationResponse(
            gross=round(gross_monthly, 2),
            net=round(gross_monthly, 2),
            breakdown=TaxBreakdown(zus=0, ppk=0, kup=0, health=0, pit=0),
        )

    if employment_type == "b2b":
        # Simplified: 19% linear tax, no ZUS/health detail
        pit = round(gross_monthly * B2B_LINEAR_RATE, 2)
        net = round(gross_monthly - pit, 2)
        return TaxCalculationResponse(
            gross=round(gross_monthly, 2),
            net=net,
            breakdown=TaxBreakdown(zus=0, ppk=0, kup=0, health=0, pit=pit),
        )

    if employment_type == "dzielo":
        # Umowa o dzieło: no ZUS, no health
        # Resolve KUP: per-income kup_type overrides global use_authors_costs
        if kup_type == "none":
            kup = 0.0
        elif kup_type == "author_50" or (kup_type is None and use_authors_costs):
            kup = min(gross_monthly * 0.5, KUP_AUTHORS_CAP_MONTHLY)
        else:
            kup = gross_monthly * 0.2
        taxable = max(gross_monthly - kup, 0)
        # Progressive PIT
        if taxable <= PIT_THRESHOLD_MONTHLY:
            pit = max(taxable * PIT_RATE_1 - TAX_FREE_MONTHLY * PIT_RATE_1, 0)
        else:
            pit_below = PIT_THRESHOLD_MONTHLY * PIT_RATE_1 - TAX_FREE_MONTHLY * PIT_RATE_1
            pit_above = (taxable - PIT_THRESHOLD_MONTHLY) * PIT_RATE_2
            pit = max(pit_below + pit_above, 0)
        pit = round(pit, 2)
        kup = round(kup, 2)
        net = round(gross_monthly - pit, 2)
        return TaxCalculationResponse(
            gross=round(gross_monthly, 2),
            net=net,
            breakdown=TaxBreakdown(zus=0, ppk=0, kup=kup, health=0, pit=pit),
        )

    # UoP and Zlecenie: ZUS + health + KUP + PIT
    zus = round(gross_monthly * ZUS_TOTAL, 2)
    ppk = round(gross_monthly * ppk_employee_rate, 2)
    base_after_zus = gross_monthly - zus

    health = round(base_after_zus * HEALTH_RATE, 2)

    # Resolve KUP: per-income kup_type overrides global use_authors_costs
    if employment_type == "uop":
        if kup_type == "none":
            kup = 0.0
        elif kup_type == "author_50" or (kup_type is None and use_authors_costs):
            kup = min(base_after_zus * 0.5, KUP_AUTHORS_CAP_MONTHLY)
        else:
            kup = KUP_STANDARD_UOP
    elif employment_type == "zlecenie":
        if kup_type == "none":
            kup = 0.0
        elif kup_type == "author_50" or (kup_type is None and use_authors_costs):
            kup = min(base_after_zus * 0.5, KUP_AUTHORS_CAP_MONTHLY)
        else:
            kup = base_after_zus * 0.2

    kup = round(kup, 2)
    taxable = max(base_after_zus - kup, 0)

    # Progressive PIT
    if taxable <= PIT_THRESHOLD_MONTHLY:
        pit = max(taxable * PIT_RATE_1 - TAX_FREE_MONTHLY * PIT_RATE_1, 0)
    else:
        pit_below = PIT_THRESHOLD_MONTHLY * PIT_RATE_1 - TAX_FREE_MONTHLY * PIT_RATE_1
        pit_above = (taxable - PIT_THRESHOLD_MONTHLY) * PIT_RATE_2
        pit = max(pit_below + pit_above, 0)

    pit = round(pit, 2)
    net = round(gross_monthly - zus - ppk - health - pit, 2)

    return TaxCalculationResponse(
        gross=round(gross_monthly, 2),
        net=net,
        breakdown=TaxBreakdown(zus=zus, ppk=ppk, kup=kup, health=health, pit=pit),
    )


@app.post("/api/tax/calculate", response_model=TaxCalculationResponse, tags=["Tax"])
async def tax_calculate(req: TaxCalculationRequest):
    """
    Stateless tax calculator: gross → net for Polish employment types.
    No authentication required.
    """
    if req.employment_type not in EMPLOYMENT_TYPES:
        raise HTTPException(status_code=400, detail=f"Invalid employment_type: {req.employment_type}")
    return calculate_net_from_gross(
        gross_monthly=req.gross_monthly,
        employment_type=req.employment_type,
        use_authors_costs=req.use_authors_costs,
        ppk_employee_rate=req.ppk_employee_rate,
        kup_type=req.kup_type,
    )


@app.get("/users/{user_id}/summary")
async def get_user_summary(
    user_id: str,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Get financial summary for the authenticated user."""
    validate_user_access(user_id, current_user)
    hid = current_user.household_id  # Use household_id for all data queries (partner sees shared data)

    try:
        print(f"[FastAPI] Getting summary for user: {hid}")
        # Get current month's start and end dates
        today = datetime.now()
        month_start = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)

        # Use MonthlyTotalsService for both income and expenses to ensure consistent
        # deduplication logic (recurring edit-duplicates, bank reconciliation, etc.)
        # Raw SUM queries here would double-count recurring entries created by edits
        # (each edit creates a new row with end_date=None, leaving two active rows).
        income_totals = MonthlyTotalsService.calculate_monthly_income(
            hid, today.year, today.month, db
        )
        monthly_income = income_totals["total"]

        expense_totals = MonthlyTotalsService.calculate_monthly_expenses(
            hid, today.year, today.month, db
        )
        monthly_expenses = expense_totals["total"]

        # Fetch monthly loan payments (exclude archived loans)
        monthly_loan_payments = db.query(func.sum(models.Loan.monthly_payment)).filter(
            models.Loan.user_id == hid,
            models.Loan.start_date <= month_end,
            (models.Loan.is_archived == False) | (models.Loan.is_archived == None)
        ).scalar() or 0

        # Fetch total loan balance (exclude archived loans)
        total_loan_balance = db.query(func.sum(models.Loan.remaining_balance)).filter(
            models.Loan.user_id == hid,
            (models.Loan.is_archived == False) | (models.Loan.is_archived == None)
        ).scalar() or 0

        # Calculate monthly balance
        monthly_balance = monthly_income - monthly_expenses - monthly_loan_payments

        # Calculate savings rate if income is not zero
        savings_rate = 0
        if monthly_income > 0:
            savings_rate = (monthly_income - monthly_expenses - monthly_loan_payments) / monthly_income

        # Calculate debt-to-income ratio if income is not zero
        debt_to_income = 0
        if monthly_income > 0:
            debt_to_income = monthly_loan_payments / monthly_income

        # Get income distribution by category using deduplicated data from MonthlyTotalsService
        income_distribution = []
        income_active = MonthlyTotalsService._query_active_income_for_month(
            hid, today.year, today.month, db
        )
        # Apply same dedup as calculate_monthly_income: filter duplicate_of_bank + recurring dedup
        income_manual = [i for i in income_active if i.bank_transaction_id is None]
        income_deduped_manual = [
            i for i in income_manual if i.reconciliation_status not in ["duplicate_of_bank"]
        ]
        _seen_inc = {}
        _non_rec_inc = []
        for i in income_deduped_manual:
            if i.is_recurring and i.end_date is None:
                key = (i.category, i.description)
                if key not in _seen_inc or i.id > _seen_inc[key].id:
                    _seen_inc[key] = i
            else:
                _non_rec_inc.append(i)
        income_deduped = (
            [i for i in income_active if i.bank_transaction_id is not None]
            + _non_rec_inc
            + list(_seen_inc.values())
        )
        income_by_category: dict = {}
        for i in income_deduped:
            income_by_category[i.category] = income_by_category.get(i.category, 0) + float(i.amount)

        total_income_dist = sum(income_by_category.values())
        for category, amount in income_by_category.items():
            percentage = (amount / total_income_dist * 100) if total_income_dist > 0 else 0
            income_distribution.append({
                "category": category,
                "amount": float(amount),
                "percentage": float(percentage)
            })

        # Get expense distribution by category using deduplicated data from MonthlyTotalsService
        expense_distribution = []
        expense_active = MonthlyTotalsService._query_active_expenses_for_month(
            hid, today.year, today.month, db
        )
        # Apply same dedup as calculate_monthly_expenses: filter duplicate_of_bank + recurring dedup
        exp_manual = [e for e in expense_active if e.bank_transaction_id is None]
        exp_deduped_manual = [
            e for e in exp_manual if e.reconciliation_status not in ["duplicate_of_bank"]
        ]
        _seen_exp = {}
        _non_rec_exp = []
        for e in exp_deduped_manual:
            if e.is_recurring and e.end_date is None:
                key = (e.category, e.description)
                if key not in _seen_exp or e.id > _seen_exp[key].id:
                    _seen_exp[key] = e
            else:
                _non_rec_exp.append(e)
        expense_deduped = (
            [e for e in expense_active if e.bank_transaction_id is not None]
            + _non_rec_exp
            + list(_seen_exp.values())
        )
        expense_by_category: dict = {}
        for e in expense_deduped:
            expense_by_category[e.category] = expense_by_category.get(e.category, 0) + float(e.amount)

        total_expense_dist = sum(expense_by_category.values())
        for category, amount in expense_by_category.items():
            percentage = (amount / total_expense_dist * 100) if total_expense_dist > 0 else 0
            expense_distribution.append({
                "category": category,
                "amount": float(amount),
                "percentage": float(percentage)
            })

        # Generate cash flow data for the rolling last 13 months (12 past + current).
        # Using a rolling window instead of "current calendar year" ensures the
        # BudgetVsActual 3-month average always has real historical data even in
        # January, and avoids polluting past-month averages with future scheduled items.
        cash_flow = []

        for offset in range(12, -1, -1):  # 12 months ago → current month (ascending)
            # Pure stdlib month arithmetic — no dateutil needed
            total_months = (today.year * 12 + today.month - 1) - offset
            m_year, m_month = divmod(total_months, 12)
            m_month += 1  # divmod gives 0-based month
            month_end_date = datetime(m_year, m_month, 1) + timedelta(days=32)
            month_end_date = month_end_date.replace(day=1) - timedelta(days=1)
            month_str = f"{m_year}-{m_month:02d}"

            # Use MonthlyTotalsService for accurate deduplication in all months
            m_income = MonthlyTotalsService.calculate_monthly_income(hid, m_year, m_month, db)
            m_expenses = MonthlyTotalsService.calculate_monthly_expenses(hid, m_year, m_month, db)

            month_income = m_income["total"]
            month_expenses = m_expenses["total"]

            # Loan payments for this month (exclude archived)
            month_loan_payments = db.query(func.sum(models.Loan.monthly_payment)).filter(
                models.Loan.user_id == hid,
                models.Loan.start_date <= month_end_date,
                (models.Loan.is_archived == False) | (models.Loan.is_archived == None)
            ).scalar() or 0

            # Calculate net flow
            net_flow = month_income - month_expenses - month_loan_payments

            cash_flow.append({
                "month": month_str,
                "income": float(month_income),
                "expenses": float(month_expenses),
                "loanPayments": float(month_loan_payments),
                "netFlow": float(net_flow),
                "year": m_year
            })

        # Fetch active loans for the user (exclude archived)
        active_loans = db.query(models.Loan).filter(
            models.Loan.user_id == hid,
            (models.Loan.is_archived == False) | (models.Loan.is_archived == None)
        ).all()
        
        # Format loans for the frontend
        loans_data = []
        for loan in active_loans:
            # Calculate progress as percentage (0-100)
            total_amount = loan.principal_amount
            paid_amount = total_amount - loan.remaining_balance
            progress = (paid_amount / total_amount * 100) if total_amount > 0 else 0

            loans_data.append({
                "id": str(loan.id),
                "description": loan.description,
                "balance": float(loan.remaining_balance),
                "monthlyPayment": float(loan.monthly_payment),
                "interestRate": float(loan.interest_rate),
                "progress": float(progress),
                "totalAmount": float(total_amount)
            })

        # Fetch savings data for the user
        # Total savings balance (all time deposits - withdrawals)
        total_deposits = db.query(func.sum(models.Saving.amount)).filter(
            models.Saving.user_id == hid,
            models.Saving.saving_type == 'deposit'
        ).scalar() or 0

        total_withdrawals = db.query(func.sum(models.Saving.amount)).filter(
            models.Saving.user_id == hid,
            models.Saving.saving_type == 'withdrawal'
        ).scalar() or 0

        total_savings_balance = total_deposits - total_withdrawals

        # Monthly savings (current month non-recurring + all recurring deposits/withdrawals)
        monthly_deposits_non_recurring = db.query(func.sum(models.Saving.amount)).filter(
            models.Saving.user_id == hid,
            models.Saving.saving_type == 'deposit',
            models.Saving.is_recurring == False,
            models.Saving.date >= month_start,
            models.Saving.date <= month_end
        ).scalar() or 0

        monthly_deposits_recurring = db.query(func.sum(models.Saving.amount)).filter(
            models.Saving.user_id == hid,
            models.Saving.saving_type == 'deposit',
            models.Saving.is_recurring == True,
            models.Saving.date <= month_end,  # Started before or during this month
            or_(
                models.Saving.end_date == None,  # No end date (ongoing)
                models.Saving.end_date >= month_start  # Or end_date is this month or later
            )
        ).scalar() or 0

        monthly_withdrawals_non_recurring = db.query(func.sum(models.Saving.amount)).filter(
            models.Saving.user_id == hid,
            models.Saving.saving_type == 'withdrawal',
            models.Saving.is_recurring == False,
            models.Saving.date >= month_start,
            models.Saving.date <= month_end
        ).scalar() or 0

        monthly_withdrawals_recurring = db.query(func.sum(models.Saving.amount)).filter(
            models.Saving.user_id == hid,
            models.Saving.saving_type == 'withdrawal',
            models.Saving.is_recurring == True,
            models.Saving.date <= month_end,  # Started before or during this month
            or_(
                models.Saving.end_date == None,  # No end date (ongoing)
                models.Saving.end_date >= month_start  # Or end_date is this month or later
            )
        ).scalar() or 0

        monthly_savings = (monthly_deposits_non_recurring + monthly_deposits_recurring) - \
                          (monthly_withdrawals_non_recurring + monthly_withdrawals_recurring)

        # Savings goals (group by category with targets)
        savings_goals = []
        savings_by_category = db.query(
            models.Saving.category,
            func.sum(
                case(
                    (models.Saving.saving_type == 'deposit', models.Saving.amount),
                    else_=-models.Saving.amount
                )
            ).label('current_amount'),
            func.max(models.Saving.target_amount).label('target_amount')
        ).filter(
            models.Saving.user_id == hid
        ).group_by(models.Saving.category).all()

        for goal in savings_by_category:
            target = float(goal.target_amount or 0)
            current = float(goal.current_amount or 0)
            # Hardcoded Baby Steps targets: emergency_fund = 1 month, six_month_fund = 6 months
            if target == 0:
                if goal.category == 'emergency_fund':
                    target = float(monthly_expenses) * 1
                elif goal.category == 'six_month_fund':
                    target = float(monthly_expenses) * 6
            # Show actual balance (no capping) — progress bar capped separately below
            effective_current = current
            progress = (effective_current / target * 100) if target > 0 else 0
            savings_goals.append({
                "category": goal.category,
                "currentAmount": effective_current,
                "targetAmount": target,
                "progress": min(float(progress), 100)
            })

        recent_activities = db.query(models.Activity).filter(
            models.Activity.user_id == hid
        ).order_by(models.Activity.timestamp.desc()).limit(20).all()

        numeric_fields = {
            'amount',
            'principal_amount',
            'remaining_balance',
            'monthly_payment',
            'target_amount',
        }

        def build_changes(activity: models.Activity) -> List[dict]:
            changes: List[dict] = []
            previous = activity.previous_values or {}
            new = activity.new_values or {}

            if activity.operation_type == 'create':
                for key, value in new.items():
                    if key in {'created_at', 'updated_at'}:
                        continue
                    if key in numeric_fields or isinstance(value, (int, float, str)):
                        changes.append({
                            'field': key,
                            'oldValue': None,
                            'newValue': value,
                        })
            elif activity.operation_type == 'delete':
                for key, value in previous.items():
                    if key in {'created_at', 'updated_at'}:
                        continue
                    if key in numeric_fields or isinstance(value, (int, float, str)):
                        changes.append({
                            'field': key,
                            'oldValue': value,
                            'newValue': None,
                        })
            else:
                for key, old_value in previous.items():
                    if key in {'created_at', 'updated_at'}:
                        continue
                    new_value = new.get(key)
                    if old_value != new_value:
                        if key in numeric_fields or isinstance(old_value, (int, float, str)) or isinstance(new_value, (int, float, str)):
                            changes.append({
                                'field': key,
                                'oldValue': old_value,
                                'newValue': new_value,
                            })
            return changes

        def activity_amount(activity: models.Activity) -> float:
            source = activity.new_values or activity.previous_values or {}
            try:
                if activity.entity_type == 'Income':
                    return float(source.get('amount', 0) or 0)
                if activity.entity_type == 'Expense':
                    return -float(source.get('amount', 0) or 0)
                if activity.entity_type == 'Loan':
                    return -float(source.get('monthly_payment', 0) or 0)
                if activity.entity_type == 'Saving':
                    amount = float(source.get('amount', 0) or 0)
                    return amount if source.get('saving_type') != 'withdrawal' else -amount
            except (TypeError, ValueError):
                pass
            return 0.0

        activities_payload = []
        for activity in recent_activities:
            base = activity.new_values or activity.previous_values or {}
            activities_payload.append({
                "id": activity.id,
                "title": base.get('description') or activity.entity_type,
                "amount": activity_amount(activity),
                "type": activity.entity_type.lower(),
                "date": activity.timestamp.isoformat() if activity.timestamp else datetime.utcnow().isoformat(),
                "operation": activity.operation_type,
                "changes": build_changes(activity),
            })

        summary = {
            "total_monthly_income": float(monthly_income),
            "total_monthly_expenses": float(monthly_expenses),
            "total_monthly_loan_payments": float(monthly_loan_payments),
            "total_loan_balance": float(total_loan_balance),
            "monthly_balance": float(monthly_balance),
            "savings_rate": float(savings_rate),
            "debt_to_income": float(debt_to_income),
            "income_distribution": income_distribution,
            "expense_distribution": expense_distribution,
            "cash_flow": cash_flow,
            "loans": loans_data,
            "activities": activities_payload,
            "total_savings_balance": float(total_savings_balance),
            "monthly_savings": float(monthly_savings),
            "savings_goals": savings_goals,
        }
        print(f"[FastAPI] Summary data: {summary}")
        return summary
    except Exception as e:
        print(f"[FastAPI] Error fetching summary: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# Activity endpoints
@app.post("/users/{user_id}/activities", response_model=Activity)
def create_activity(
    user_id: str,
    activity: ActivityCreate,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Create an activity log for the authenticated user."""
    validate_user_access(user_id, current_user)

    try:
        print(f"[FastAPI] Creating activity for user: {current_user.household_id}")
        db_activity = models.Activity(
            user_id=current_user.household_id,
            entity_type=activity.entity_type,
            operation_type=activity.operation_type,
            entity_id=activity.entity_id,
            previous_values=activity.previous_values,
            new_values=activity.new_values
        )
        db.add(db_activity)
        db.commit()
        db.refresh(db_activity)
        print(f"[FastAPI] Created activity: {db_activity}")
        return db_activity
    except HTTPException:
        raise
    except Exception as e:
        print(f"[FastAPI] Error in create_activity: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/users/{user_id}/activities", response_model=List[Activity])
def get_user_activities(
    user_id: str,
    skip: int = 0,
    limit: int = 50,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Get activity log for the authenticated user."""
    validate_user_access(user_id, current_user)

    try:
        print(f"[FastAPI] Getting activities for user: {current_user.household_id}")
        activities = db.query(models.Activity)\
            .filter(models.Activity.user_id == current_user.household_id)\
            .order_by(models.Activity.timestamp.desc())\
            .offset(skip)\
            .limit(limit)\
            .all()
        print(f"[FastAPI] Found {len(activities)} activities")
        return activities
    except HTTPException:
        raise
    except Exception as e:
        print(f"[FastAPI] Error in get_user_activities: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/reports/yearly-budget")
def get_yearly_budget(
    start_date: str,
    end_date: str,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Get yearly budget report for the authenticated user."""
    try:
        print(f"[FastAPI] Getting budget for user: {current_user.household_id}, period: {start_date} to {end_date}")
        
        try:
            # Parse date strings to datetime objects
            start_datetime = datetime.strptime(start_date, "%Y-%m")
            end_datetime = datetime.strptime(end_date, "%Y-%m")
            
            # Get the last day of the end month
            end_datetime = end_datetime.replace(
                day=calendar.monthrange(end_datetime.year, end_datetime.month)[1],
                hour=23, minute=59, second=59
            )
            
            print(f"[FastAPI] Parsed dates - Start: {start_datetime}, End: {end_datetime}")
        except ValueError as e:
            print(f"[FastAPI] Error parsing dates: {str(e)}")
            raise HTTPException(status_code=400, detail=f"Invalid date format: {str(e)}")
        
        # Get all non-recurring incomes for the period
        try:
            regular_incomes = db.query(models.Income).filter(
                models.Income.user_id == current_user.household_id,
                models.Income.date >= start_datetime,
                models.Income.date <= end_datetime,
                models.Income.is_recurring == False
            ).all()
            print(f"[FastAPI] Found {len(regular_incomes)} regular incomes")
        except Exception as e:
            print(f"[FastAPI] Error fetching regular incomes: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Error fetching regular incomes: {str(e)}")

        # Get recurring incomes
        try:
            recurring_incomes = db.query(models.Income).filter(
                models.Income.user_id == current_user.household_id,
                models.Income.is_recurring == True,
                models.Income.date <= end_datetime
            ).all()
            print(f"[FastAPI] Found {len(recurring_incomes)} recurring incomes")
        except Exception as e:
            print(f"[FastAPI] Error fetching recurring incomes: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Error fetching recurring incomes: {str(e)}")

        # Get all non-recurring expenses for the period
        try:
            regular_expenses = db.query(models.Expense).filter(
                models.Expense.user_id == current_user.household_id,
                models.Expense.date >= start_datetime,
                models.Expense.date <= end_datetime,
                models.Expense.is_recurring == False
            ).all()
            print(f"[FastAPI] Found {len(regular_expenses)} regular expenses")
        except Exception as e:
            print(f"[FastAPI] Error fetching regular expenses: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Error fetching regular expenses: {str(e)}")

        # Get recurring expenses
        try:
            recurring_expenses = db.query(models.Expense).filter(
                models.Expense.user_id == current_user.household_id,
                models.Expense.is_recurring == True,
                models.Expense.date <= end_datetime
            ).all()
            print(f"[FastAPI] Found {len(recurring_expenses)} recurring expenses")
        except Exception as e:
            print(f"[FastAPI] Error fetching recurring expenses: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Error fetching recurring expenses: {str(e)}")

        # Get all loans that could be active during this period
        try:
            loans = db.query(models.Loan).filter(
                models.Loan.user_id == current_user.household_id,
                models.Loan.start_date <= end_datetime
            ).all()
            print(f"[FastAPI] Found {len(loans)} loans")
            for loan in loans:
                print(f"[FastAPI] Loan: id={loan.id}, monthly_payment={loan.monthly_payment}, start_date={loan.start_date}, term_months={loan.term_months}")
        except Exception as e:
            print(f"[FastAPI] Error fetching loans: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Error fetching loans: {str(e)}")

        # Get all non-recurring savings for the period
        try:
            regular_savings = db.query(models.Saving).filter(
                models.Saving.user_id == current_user.household_id,
                models.Saving.date >= start_datetime,
                models.Saving.date <= end_datetime,
                models.Saving.is_recurring == False
            ).all()
            print(f"[FastAPI] Found {len(regular_savings)} regular savings")
        except Exception as e:
            print(f"[FastAPI] Error fetching regular savings: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Error fetching regular savings: {str(e)}")

        # Get recurring savings
        try:
            recurring_savings = db.query(models.Saving).filter(
                models.Saving.user_id == current_user.household_id,
                models.Saving.is_recurring == True,
                models.Saving.date <= end_datetime
            ).all()
            print(f"[FastAPI] Found {len(recurring_savings)} recurring savings")
        except Exception as e:
            print(f"[FastAPI] Error fetching recurring savings: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Error fetching recurring savings: {str(e)}")

        # Initialize monthly data
        monthly_data = {}
        
        # Generate a list of all months in the range
        current_date = start_datetime
        while current_date <= end_datetime:
            try:
                month_name = current_date.strftime("%B")
                month_idx = current_date.month
                year = current_date.year

                print(f"[FastAPI] Processing month: {month_name} year: {year}")

                # Convert current_date to date for comparison
                current_date_only = current_date.date()

                # Regular incomes for this month
                monthly_regular_incomes = [
                    {
                        "id": inc.id,
                        "title": inc.description,
                        "amount": float(inc.amount),
                        "category": inc.category,
                        "date": inc.date.isoformat()
                    }
                    for inc in regular_incomes 
                    if inc.date.year == year and inc.date.month == month_idx
                ]
                
                # Add recurring incomes if they started before or during this month
                # and haven't ended yet (end_date is None or >= current month)
                monthly_recurring_incomes = [
                    {
                        "id": inc.id,
                        "title": inc.description,
                        "amount": float(inc.amount),
                        "category": inc.category,
                        "date": inc.date.isoformat()
                    }
                    for inc in recurring_incomes
                    if inc.date <= current_date.date() and (
                        inc.end_date is None or inc.end_date >= current_date_only
                    )
                ]
                
                # Regular expenses for this month
                monthly_regular_expenses = [
                    {
                        "id": exp.id,
                        "title": exp.description,
                        "amount": float(exp.amount),
                        "category": exp.category,
                        "date": exp.date.isoformat()
                    }
                    for exp in regular_expenses 
                    if exp.date.year == year and exp.date.month == month_idx
                ]
                
                # Add recurring expenses if they started before or during this month
                # and haven't ended yet (end_date is None or >= current month)
                monthly_recurring_expenses = [
                    {
                        "id": exp.id,
                        "title": exp.description,
                        "amount": float(exp.amount),
                        "category": exp.category,
                        "date": exp.date.isoformat()
                    }
                    for exp in recurring_expenses
                    if exp.date <= current_date.date() and (
                        exp.end_date is None or exp.end_date >= current_date_only
                    )
                ]

                # Get active loans for this month
                monthly_loans = []
                monthly_loan_payments_total = 0.0
                
                for loan in loans:
                    try:
                        # Convert loan.start_date to datetime for comparison
                        loan_start = datetime.combine(loan.start_date, datetime.min.time())

                        # Calculate loan end date based on start date and term
                        loan_end_date = loan_start + timedelta(days=30 * loan.term_months)
                        
                        # Check if loan is active in this month
                        if loan_start.date() <= current_date_only and loan_end_date.date() >= current_date_only:
                            loan_payment = float(loan.monthly_payment) if loan.monthly_payment is not None else 0.0
                            monthly_loan_payments_total += loan_payment
                            
                            monthly_loans.append({
                                "id": loan.id,
                                "loan_type": loan.loan_type,
                                "description": loan.description,
                                "principal_amount": float(loan.principal_amount) if loan.principal_amount is not None else 0.0,
                                "remaining_balance": float(loan.remaining_balance) if loan.remaining_balance is not None else 0.0,
                                "interest_rate": float(loan.interest_rate) if loan.interest_rate is not None else 0.0,
                                "monthly_payment": loan_payment,
                                "start_date": loan.start_date.isoformat(),
                                "term_months": loan.term_months
                            })
                    except (ValueError, TypeError, AttributeError) as e:
                        print(f"[FastAPI] Error processing loan {loan.id}: {str(e)}")
                        continue

                # Regular savings for this month
                monthly_regular_savings = [
                    {
                        "id": sav.id,
                        "title": sav.description,
                        "amount": float(sav.amount),
                        "category": sav.category,
                        "saving_type": sav.saving_type,
                        "date": sav.date.isoformat()
                    }
                    for sav in regular_savings
                    if sav.date.year == year and sav.date.month == month_idx
                ]

                # Add recurring savings if they started before or during this month
                # and haven't ended yet (end_date is None or >= current month)
                monthly_recurring_savings = [
                    {
                        "id": sav.id,
                        "title": sav.description,
                        "amount": float(sav.amount),
                        "category": sav.category,
                        "saving_type": sav.saving_type,
                        "date": sav.date.isoformat()
                    }
                    for sav in recurring_savings
                    if sav.date <= current_date.date() and (
                        sav.end_date is None or sav.end_date >= current_date_only
                    )
                ]

                # Combine all savings for this month
                all_monthly_savings = monthly_regular_savings + monthly_recurring_savings

                # Calculate savings totals (deposits - withdrawals)
                total_deposits = sum(
                    sav["amount"] for sav in all_monthly_savings
                    if sav["saving_type"] == "deposit"
                )
                total_withdrawals = sum(
                    sav["amount"] for sav in all_monthly_savings
                    if sav["saving_type"] == "withdrawal"
                )
                net_savings = float(total_deposits - total_withdrawals)

                # Calculate totals
                total_regular_income = sum(inc["amount"] for inc in monthly_regular_incomes)
                total_recurring_income = sum(inc["amount"] for inc in monthly_recurring_incomes)
                total_regular_expenses = sum(exp["amount"] for exp in monthly_regular_expenses)
                total_recurring_expenses = sum(exp["amount"] for exp in monthly_recurring_expenses)
                
                total_income = float(total_regular_income + total_recurring_income)
                total_expenses = float(total_regular_expenses + total_recurring_expenses)
                
                print(f"[FastAPI] {month_name} {year} - Income: {total_income}, Expenses: {total_expenses}, Loan Payments: {monthly_loan_payments_total}, Savings: {net_savings}")

                monthly_data[f"{month_name} {year}"] = {
                    "incomes": monthly_regular_incomes + monthly_recurring_incomes,
                    "expenses": monthly_regular_expenses + monthly_recurring_expenses,
                    "loanPayments": monthly_loans,
                    "savings": all_monthly_savings,
                    "totals": {
                        "income": total_income,
                        "expenses": total_expenses,
                        "loanPayments": monthly_loan_payments_total,
                        "savings": net_savings,
                        "deposits": float(total_deposits),
                        "withdrawals": float(total_withdrawals),
                        "balance": total_income - total_expenses - monthly_loan_payments_total
                    }
                }
                
                # Move to next month
                if month_idx == 12:
                    current_date = current_date.replace(year=year + 1, month=1)
                else:
                    current_date = current_date.replace(month=month_idx + 1)
                    
            except Exception as e:
                print(f"[FastAPI] Error processing month {month_name} {year}: {str(e)}")
                continue
        
        print(f"[FastAPI] Successfully generated budget data for period")
        return monthly_data
        
    except Exception as e:
        print(f"[FastAPI] Error in get_yearly_budget: {str(e)}")
        import traceback
        print(f"[FastAPI] Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/users/{user_id}/export")
async def export_user_data(
    user_id: str,
    format: str = Query(..., description="Export format: 'json', 'csv', or 'xlsx'"),
    save_backup: bool = Query(False, description="Whether to save a backup on the server (JSON only)"),
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Export user data in various formats for the authenticated user."""
    validate_user_access(user_id, current_user)

    try:
        # Check subscription for export format
        can_export, message = SubscriptionService.can_export_format(current_user.household_id, format, db)
        if not can_export:
            raise HTTPException(status_code=403, detail=message)

        # Fetch all user data
        settings = db.query(models.Settings).filter(models.Settings.user_id == current_user.household_id).first()
        expenses = db.query(models.Expense).filter(models.Expense.user_id == current_user.household_id).all()
        incomes = db.query(models.Income).filter(models.Income.user_id == current_user.household_id).all()
        loans = db.query(models.Loan).filter(models.Loan.user_id == current_user.household_id).all()
        savings = db.query(models.Saving).filter(models.Saving.user_id == current_user.household_id).all()

        # Fetch additional data
        loan_payments = db.query(models.LoanPayment).filter(models.LoanPayment.user_id == current_user.household_id).all()
        savings_goals = db.query(models.SavingsGoal).filter(models.SavingsGoal.user_id == current_user.household_id).all()

        # Build goal_id -> goal name lookup for savings
        goal_names = {g.id: g.name for g in savings_goals}

        # Prepare data structure for JSON
        data = {
            "settings": {
                "language": settings.language if settings else "en",
                "currency": settings.currency if settings else "USD",
                "emergency_fund_target": settings.emergency_fund_target if settings else 1000,
                "emergency_fund_months": settings.emergency_fund_months if settings else 3,
                "base_currency": settings.base_currency if settings else "USD",
                "employment_status": settings.employment_status if settings else None,
                "tax_form": settings.tax_form if settings else None,
                "birth_year": settings.birth_year if settings else None,
                "use_authors_costs": settings.use_authors_costs if settings else False,
                "ppk_enrolled": settings.ppk_enrolled if settings else None,
                "ppk_employee_rate": settings.ppk_employee_rate if settings else None,
                "ppk_employer_rate": settings.ppk_employer_rate if settings else None,
                "children_count": settings.children_count if settings else 0,
            },
            "expenses": [
                {
                    "category": expense.category,
                    "description": expense.description,
                    "amount": expense.amount,
                    "date": expense.date.isoformat(),
                    "is_recurring": expense.is_recurring,
                    "end_date": expense.end_date.isoformat() if expense.end_date else None,
                }
                for expense in expenses
            ],
            "incomes": [
                {
                    "category": income.category,
                    "description": income.description,
                    "amount": income.amount,
                    "date": income.date.isoformat(),
                    "is_recurring": income.is_recurring,
                    "end_date": income.end_date.isoformat() if income.end_date else None,
                    "employment_type": income.employment_type,
                    "gross_amount": income.gross_amount,
                    "is_gross": income.is_gross,
                }
                for income in incomes
            ],
            "loans": [
                {
                    "loan_type": loan.loan_type,
                    "description": loan.description,
                    "principal_amount": loan.principal_amount,
                    "remaining_balance": loan.remaining_balance,
                    "interest_rate": loan.interest_rate,
                    "monthly_payment": loan.monthly_payment,
                    "term_months": loan.term_months,
                    "start_date": loan.start_date.isoformat(),
                    "due_day": loan.due_day,
                    "overpayment_fee_percent": loan.overpayment_fee_percent,
                    "overpayment_fee_waived_until": loan.overpayment_fee_waived_until.isoformat() if loan.overpayment_fee_waived_until else None,
                    "is_archived": loan.is_archived,
                    "payments": [
                        {
                            "amount": p.amount,
                            "payment_date": p.payment_date.isoformat(),
                            "payment_type": p.payment_type,
                            "covers_month": p.covers_month,
                            "covers_year": p.covers_year,
                            "notes": p.notes,
                        }
                        for p in loan_payments if p.loan_id == loan.id
                    ],
                }
                for loan in loans
            ],
            "savings": [
                {
                    "category": saving.category,
                    "description": saving.description,
                    "amount": saving.amount,
                    "date": saving.date.isoformat(),
                    "is_recurring": saving.is_recurring,
                    "end_date": saving.end_date.isoformat() if saving.end_date else None,
                    "target_amount": saving.target_amount,
                    "saving_type": saving.saving_type,
                    "account_type": saving.account_type,
                    "annual_return_rate": saving.annual_return_rate,
                    "goal_name": goal_names.get(saving.goal_id) if saving.goal_id else None,
                }
                for saving in savings
            ],
            "savings_goals": [
                {
                    "name": goal.name,
                    "category": goal.category,
                    "target_amount": goal.target_amount,
                    "current_amount": goal.current_amount,
                    "deadline": goal.deadline.isoformat() if goal.deadline else None,
                    "icon": goal.icon,
                    "color": goal.color,
                    "status": goal.status,
                    "priority": goal.priority,
                    "notes": goal.notes,
                }
                for goal in savings_goals
            ],
        }

        if format.lower() == 'json':
            # Save backup to database if requested
            if save_backup:
                import json
                data_json = json.dumps(data)
                filename = f"home_budget_export_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                backup = models.DataExportBackup(
                    user_id=current_user.household_id,
                    data=data,
                    format="json",
                    filename=filename,
                    size_bytes=len(data_json.encode('utf-8'))
                )
                db.add(backup)
                db.commit()
            return JSONResponse(content=data)
        elif format.lower() == 'csv':
            # Create CSV buffer
            output = io.StringIO()
            writer = csv.writer(output)

            # Write headers
            writer.writerow(['Type', 'Category/Loan Type', 'Description', 'Amount/Principal', 'Date/Start Date', 'Is Recurring/Interest Rate', 'Target Amount/Remaining Balance', 'Monthly Payment/Saving Type', 'Term Months'])

            # Write settings
            writer.writerow(['Settings', 'language', settings.language if settings else "en", '', '', '', '', '', ''])
            writer.writerow(['Settings', 'currency', settings.currency if settings else "USD", '', '', '', '', '', ''])
            writer.writerow(['Settings', 'emergency_fund_target', settings.emergency_fund_target if settings else 1000, '', '', '', '', '', ''])
            writer.writerow(['Settings', 'emergency_fund_months', settings.emergency_fund_months if settings else 3, '', '', '', '', '', ''])

            # Write expenses
            for expense in expenses:
                writer.writerow([
                    'Expense',
                    expense.category,
                    expense.description,
                    expense.amount,
                    expense.date.isoformat(),
                    expense.is_recurring,
                    '', '', ''
                ])

            # Write incomes
            for income in incomes:
                writer.writerow([
                    'Income',
                    income.category,
                    income.description,
                    income.amount,
                    income.date.isoformat(),
                    income.is_recurring,
                    '', '', ''
                ])

            # Write loans
            for loan in loans:
                writer.writerow([
                    'Loan',
                    loan.loan_type,
                    loan.description,
                    loan.principal_amount,
                    loan.start_date.isoformat(),
                    loan.interest_rate,
                    loan.remaining_balance,
                    loan.monthly_payment,
                    loan.term_months
                ])
                
            # Write savings
            for saving in savings:
                writer.writerow([
                    'Saving',
                    saving.category,
                    saving.description,
                    saving.amount,
                    saving.date.isoformat(),
                    saving.is_recurring,
                    saving.target_amount or '',
                    saving.saving_type,
                    ''
                ])

            # Prepare the response
            output.seek(0)
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="text/csv",
                headers={
                    "Content-Disposition": f"attachment; filename=home_budget_export_{user_id}.csv"
                }
            )
        elif format.lower() == 'xlsx':
            # Create Excel workbook
            wb = Workbook()
            
            # Create styles
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')

            def style_header_row(worksheet):
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill

            # Settings sheet
            settings_sheet = wb.active
            settings_sheet.title = 'Settings'
            settings_sheet.append(['Setting', 'Value'])
            settings_sheet.append(['Language', settings.language if settings else "en"])
            settings_sheet.append(['Currency', settings.currency if settings else "USD"])
            settings_sheet.append(['Emergency Fund Target', settings.emergency_fund_target if settings else 1000])
            settings_sheet.append(['Emergency Fund Months', settings.emergency_fund_months if settings else 3])
            style_header_row(settings_sheet)
            
            # Expenses sheet
            expenses_sheet = wb.create_sheet('Expenses')
            expenses_sheet.append(['Category', 'Description', 'Amount', 'Date', 'Is Recurring'])
            for expense in expenses:
                expenses_sheet.append([
                    expense.category,
                    expense.description,
                    expense.amount,
                    expense.date.isoformat(),
                    expense.is_recurring
                ])
            style_header_row(expenses_sheet)

            # Income sheet
            income_sheet = wb.create_sheet('Income')
            income_sheet.append(['Category', 'Description', 'Amount', 'Date', 'Is Recurring'])
            for income in incomes:
                income_sheet.append([
                    income.category,
                    income.description,
                    income.amount,
                    income.date.isoformat(),
                    income.is_recurring
                ])
            style_header_row(income_sheet)

            # Loans sheet
            loans_sheet = wb.create_sheet('Loans')
            loans_sheet.append([
                'Type', 'Description', 'Principal Amount', 'Remaining Balance',
                'Interest Rate', 'Monthly Payment', 'Term Months', 'Start Date'
            ])
            for loan in loans:
                loans_sheet.append([
                    loan.loan_type,
                    loan.description,
                    loan.principal_amount,
                    loan.remaining_balance,
                    loan.interest_rate,
                    loan.monthly_payment,
                    loan.term_months,
                    loan.start_date.isoformat()
                ])
            style_header_row(loans_sheet)
            
            # Savings sheet
            savings_sheet = wb.create_sheet('Savings')
            savings_sheet.append([
                'Category', 'Description', 'Amount', 'Date', 'Is Recurring',
                'Target Amount', 'Saving Type'
            ])
            for saving in savings:
                savings_sheet.append([
                    saving.category,
                    saving.description,
                    saving.amount,
                    saving.date.isoformat(),
                    saving.is_recurring,
                    saving.target_amount,
                    saving.saving_type
                ])
            style_header_row(savings_sheet)

            # Adjust column widths
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for column in ws.columns:
                    max_length = 0
                    column = list(column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width

            # Save to buffer
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            # Return the Excel file
            return StreamingResponse(
                iter([excel_buffer.getvalue()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=home_budget_export_{user_id}.xlsx"
                }
            )
        else:
            raise HTTPException(status_code=400, detail="Invalid format. Use 'json', 'csv', or 'xlsx'")

    except Exception as e:
        print(f"[FastAPI] Error in export_user_data: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/users/{user_id}/export-backups")
async def list_export_backups(
    user_id: str,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """List all export backups for the authenticated user."""
    validate_user_access(user_id, current_user)

    backups = db.query(models.DataExportBackup).filter(
        models.DataExportBackup.user_id == current_user.household_id
    ).order_by(models.DataExportBackup.created_at.desc()).all()

    return [
        {
            "id": backup.id,
            "filename": backup.filename,
            "format": backup.format,
            "size_bytes": backup.size_bytes,
            "created_at": backup.created_at.isoformat() if backup.created_at else None
        }
        for backup in backups
    ]


@app.get("/users/{user_id}/export-backups/{backup_id}")
async def download_export_backup(
    user_id: str,
    backup_id: int,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Download a specific export backup."""
    validate_user_access(user_id, current_user)

    backup = db.query(models.DataExportBackup).filter(
        models.DataExportBackup.id == backup_id,
        models.DataExportBackup.user_id == current_user.household_id
    ).first()

    if not backup:
        raise HTTPException(status_code=404, detail="Backup not found")

    return JSONResponse(
        content=backup.data,
        headers={
            "Content-Disposition": f"attachment; filename={backup.filename}"
        }
    )


@app.delete("/users/{user_id}/export-backups/{backup_id}")
async def delete_export_backup(
    user_id: str,
    backup_id: int,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Delete a specific export backup."""
    validate_user_access(user_id, current_user)

    backup = db.query(models.DataExportBackup).filter(
        models.DataExportBackup.id == backup_id,
        models.DataExportBackup.user_id == current_user.household_id
    ).first()

    if not backup:
        raise HTTPException(status_code=404, detail="Backup not found")

    db.delete(backup)
    db.commit()

    return {"message": "Backup deleted successfully"}


@app.post("/users/{user_id}/import")
async def import_user_data(
    user_id: str,
    payload: ImportPayload,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Import user data for the authenticated user."""
    validate_user_access(user_id, current_user)

    try:
        data = payload.data
        clear_existing = payload.clear_existing
        def serialize_expense(expense: models.Expense) -> dict:
            return {
                "category": expense.category,
                "description": expense.description,
                "amount": expense.amount,
                "is_recurring": expense.is_recurring,
                "date": expense.date.isoformat() if expense.date else None,
            }

        def serialize_income(income: models.Income) -> dict:
            return {
                "category": income.category,
                "description": income.description,
                "amount": income.amount,
                "is_recurring": income.is_recurring,
                "date": income.date.isoformat() if income.date else None,
            }

        def serialize_loan(loan: models.Loan) -> dict:
            return {
                "loan_type": loan.loan_type,
                "description": loan.description,
                "principal_amount": loan.principal_amount,
                "remaining_balance": loan.remaining_balance,
                "interest_rate": loan.interest_rate,
                "monthly_payment": loan.monthly_payment,
                "term_months": loan.term_months,
                "start_date": loan.start_date.isoformat() if loan.start_date else None,
            }

        def serialize_saving(saving: models.Saving) -> dict:
            return {
                "category": saving.category,
                "description": saving.description,
                "amount": saving.amount,
                "is_recurring": saving.is_recurring,
                "date": saving.date.isoformat() if saving.date else None,
                "target_amount": saving.target_amount,
                "saving_type": saving.saving_type,
            }

        def add_activity(
            entity_type: str,
            operation: str,
            entity_id: int,
            previous_values: Optional[dict] = None,
            new_values: Optional[dict] = None,
        ):
            activity = models.Activity(
                user_id=current_user.household_id,
                entity_type=entity_type,
                operation_type=operation,
                entity_id=entity_id,
                previous_values=previous_values,
                new_values=new_values,
            )
            db.add(activity)

        # Clear existing data if requested
        if clear_existing:
            existing_expenses = db.query(models.Expense).filter(models.Expense.user_id == current_user.household_id).all()
            for expense in existing_expenses:
                add_activity(
                    entity_type="Expense",
                    operation="delete",
                    entity_id=expense.id,
                    previous_values=serialize_expense(expense),
                )
                db.delete(expense)

            existing_incomes = db.query(models.Income).filter(models.Income.user_id == current_user.household_id).all()
            for income in existing_incomes:
                add_activity(
                    entity_type="Income",
                    operation="delete",
                    entity_id=income.id,
                    previous_values=serialize_income(income),
                )
                db.delete(income)

            existing_loans = db.query(models.Loan).filter(models.Loan.user_id == current_user.household_id).all()
            for loan in existing_loans:
                add_activity(
                    entity_type="Loan",
                    operation="delete",
                    entity_id=loan.id,
                    previous_values=serialize_loan(loan),
                )
                db.delete(loan)

            existing_savings = db.query(models.Saving).filter(models.Saving.user_id == current_user.household_id).all()
            for saving in existing_savings:
                add_activity(
                    entity_type="Saving",
                    operation="delete",
                    entity_id=saving.id,
                    previous_values=serialize_saving(saving),
                )
                db.delete(saving)

            # Clear savings goals
            existing_goals = db.query(models.SavingsGoal).filter(models.SavingsGoal.user_id == current_user.household_id).all()
            for goal in existing_goals:
                db.delete(goal)

            db.commit()
            print(f"[FastAPI] Cleared existing data for user: {current_user.household_id}")

        # Import settings
        if "settings" in data:
            settings = db.query(models.Settings).filter(models.Settings.user_id == current_user.household_id).first()
            if settings:
                s = data["settings"]
                settings.language = s.get("language", settings.language)
                settings.currency = s.get("currency", settings.currency)
                settings.ai = s.get("ai", settings.ai)
                settings.emergency_fund_target = s.get("emergency_fund_target", settings.emergency_fund_target)
                settings.emergency_fund_months = s.get("emergency_fund_months", settings.emergency_fund_months)
                settings.base_currency = s.get("base_currency", settings.base_currency)
                settings.employment_status = s.get("employment_status", settings.employment_status)
                settings.tax_form = s.get("tax_form", settings.tax_form)
                settings.birth_year = s.get("birth_year", settings.birth_year)
                settings.use_authors_costs = s.get("use_authors_costs", settings.use_authors_costs)
                settings.ppk_enrolled = s.get("ppk_enrolled", settings.ppk_enrolled)
                settings.ppk_employee_rate = s.get("ppk_employee_rate", settings.ppk_employee_rate)
                settings.ppk_employer_rate = s.get("ppk_employer_rate", settings.ppk_employer_rate)
                settings.children_count = s.get("children_count", settings.children_count)
                db.commit()

        # Import expenses
        if "expenses" in data:
            for expense_data in data["expenses"]:
                expense = models.Expense(
                    user_id=current_user.household_id,
                    category=expense_data["category"],
                    description=expense_data["description"],
                    amount=expense_data["amount"],
                    date=datetime.fromisoformat(expense_data["date"].replace("Z", "+00:00")),
                    is_recurring=expense_data.get("is_recurring", False),
                    end_date=datetime.fromisoformat(expense_data["end_date"].replace("Z", "+00:00")) if expense_data.get("end_date") else None,
                )
                db.add(expense)
                db.flush()
                add_activity(
                    entity_type="Expense",
                    operation="create",
                    entity_id=expense.id,
                    new_values=serialize_expense(expense),
                )

        # Import incomes
        if "incomes" in data:
            for income_data in data["incomes"]:
                income = models.Income(
                    user_id=current_user.household_id,
                    category=income_data["category"],
                    description=income_data["description"],
                    amount=income_data["amount"],
                    date=datetime.fromisoformat(income_data["date"].replace("Z", "+00:00")),
                    is_recurring=income_data.get("is_recurring", False),
                    end_date=datetime.fromisoformat(income_data["end_date"].replace("Z", "+00:00")) if income_data.get("end_date") else None,
                    employment_type=income_data.get("employment_type"),
                    gross_amount=income_data.get("gross_amount"),
                    is_gross=income_data.get("is_gross", False),
                )
                db.add(income)
                db.flush()
                add_activity(
                    entity_type="Income",
                    operation="create",
                    entity_id=income.id,
                    new_values=serialize_income(income),
                )

        # Import loans
        if "loans" in data:
            for loan_data in data["loans"]:
                loan = models.Loan(
                    user_id=current_user.household_id,
                    loan_type=loan_data["loan_type"],
                    description=loan_data["description"],
                    principal_amount=loan_data["principal_amount"],
                    remaining_balance=loan_data["remaining_balance"],
                    interest_rate=loan_data["interest_rate"],
                    monthly_payment=loan_data["monthly_payment"],
                    term_months=loan_data["term_months"],
                    start_date=datetime.fromisoformat(loan_data["start_date"].replace("Z", "+00:00")),
                    due_day=loan_data.get("due_day", 1),
                    overpayment_fee_percent=loan_data.get("overpayment_fee_percent", 0),
                    overpayment_fee_waived_until=datetime.fromisoformat(loan_data["overpayment_fee_waived_until"].replace("Z", "+00:00")) if loan_data.get("overpayment_fee_waived_until") else None,
                    is_archived=loan_data.get("is_archived", False),
                )
                db.add(loan)
                db.flush()
                add_activity(
                    entity_type="Loan",
                    operation="create",
                    entity_id=loan.id,
                    new_values=serialize_loan(loan),
                )
                # Import loan payments
                for payment_data in loan_data.get("payments", []):
                    payment = models.LoanPayment(
                        loan_id=loan.id,
                        user_id=current_user.household_id,
                        amount=payment_data["amount"],
                        payment_date=datetime.fromisoformat(payment_data["payment_date"].replace("Z", "+00:00")),
                        payment_type=payment_data["payment_type"],
                        covers_month=payment_data.get("covers_month"),
                        covers_year=payment_data.get("covers_year"),
                        notes=payment_data.get("notes"),
                    )
                    db.add(payment)

        # Import savings
        if "savings" in data:
            for saving_data in data["savings"]:
                saving = models.Saving(
                    user_id=current_user.household_id,
                    category=saving_data["category"],
                    description=saving_data["description"],
                    amount=saving_data["amount"],
                    date=datetime.fromisoformat(saving_data["date"].replace("Z", "+00:00")),
                    is_recurring=saving_data.get("is_recurring", False),
                    end_date=datetime.fromisoformat(saving_data["end_date"].replace("Z", "+00:00")) if saving_data.get("end_date") else None,
                    target_amount=saving_data.get("target_amount"),
                    saving_type=saving_data["saving_type"],
                    account_type=saving_data.get("account_type", "standard"),
                    annual_return_rate=saving_data.get("annual_return_rate"),
                )
                db.add(saving)
                db.flush()
                add_activity(
                    entity_type="Saving",
                    operation="create",
                    entity_id=saving.id,
                    new_values=serialize_saving(saving),
                )

        # Import savings goals
        if "savings_goals" in data:
            for goal_data in data["savings_goals"]:
                goal = models.SavingsGoal(
                    user_id=current_user.household_id,
                    name=goal_data["name"],
                    category=goal_data["category"],
                    target_amount=goal_data["target_amount"],
                    current_amount=goal_data.get("current_amount", 0),
                    deadline=datetime.fromisoformat(goal_data["deadline"].replace("Z", "+00:00")) if goal_data.get("deadline") else None,
                    icon=goal_data.get("icon"),
                    color=goal_data.get("color"),
                    status=goal_data.get("status", "active"),
                    priority=goal_data.get("priority", 0),
                    notes=goal_data.get("notes"),
                )
                db.add(goal)

        db.commit()
        return {"message": "Data imported successfully"}

    except Exception as e:
        db.rollback()
        print(f"[FastAPI] Error in import_user_data: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# Insights endpoints
@app.get("/users/{user_id}/insights", response_model=InsightsResponse)
async def get_user_insights(
    user_id: str,
    refresh: bool = False,  # Changed default to False
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Get AI insights for the authenticated user."""
    validate_user_access(user_id, current_user)

    try:
        # Check for server-side OpenAI API key
        if not os.environ.get("OPENAI_API_KEY"):
            raise HTTPException(status_code=503, detail="API_KEY_MISSING")

        # If manual refresh requested, bypass cache
        if refresh:
            return await _refresh_insights_internal(current_user.household_id, db)

        # Get current financial data
        current_data = await get_user_financial_data(current_user.household_id, db)

        # Get user's language preference
        language = current_data.get("settings", {}).get("language", "en")

        # Calculate current totals
        current_income = sum(income["amount"] for income in current_data["incomes"])
        current_expenses = sum(expense["amount"] for expense in current_data["expenses"])
        current_loans = sum(loan["remaining_balance"] for loan in current_data["loans"])

        # Check for valid cache for the current language
        cache = db.query(models.InsightsCache).filter(
            models.InsightsCache.user_id == current_user.household_id,
            models.InsightsCache.language == language,
            models.InsightsCache.is_stale == False
        ).order_by(models.InsightsCache.created_at.desc()).first()

        if cache:
            # Cache exists — always return it (user must manually refresh via button)
            return {
                **cache.insights,
                "metadata": {
                    "isCached": True,
                    "createdAt": cache.created_at.isoformat(),
                    "lastRefreshDate": cache.last_refresh_date.isoformat(),
                    "language": language,
                    "validityReason": "Using cached insights",
                }
            }

        # No valid cache exists for this language, generate new insights
        return await _refresh_insights_internal(current_user.household_id, db)
    except HTTPException:
        raise
    except Exception as e:
        print(f"[FastAPI] Error in get_user_insights: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/users/{user_id}/insights/refresh", response_model=InsightsResponse)
async def refresh_user_insights(
    user_id: str,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db)
):
    """Refresh AI insights for the authenticated user."""
    validate_user_access(user_id, current_user)

    return await _refresh_insights_internal(current_user.household_id, db)


async def _refresh_insights_internal(user_id: str, db: Session):
    """Internal function to refresh insights for a user."""
    try:
        # Check for server-side OpenAI API key
        api_key = os.environ.get("OPENAI_API_KEY")
        if not api_key:
            raise HTTPException(status_code=503, detail="API_KEY_MISSING")

        # Get user's financial data
        user_data = await get_user_financial_data(user_id, db)

        # Get user's language preference
        language = user_data.get("settings", {}).get("language", "en")

        # Calculate totals for cache comparison
        total_income = sum(income["amount"] for income in user_data["incomes"])
        total_expenses = sum(expense["amount"] for expense in user_data["expenses"])
        total_loans = sum(loan["remaining_balance"] for loan in user_data["loans"])

        # Generate new insights
        insights = await generate_insights(user_data, api_key)

        # Mark existing cache entries as stale for this language
        db.query(models.InsightsCache).filter(
            models.InsightsCache.user_id == user_id,
            models.InsightsCache.language == language,
            models.InsightsCache.is_stale == False
        ).update({"is_stale": True})

        # Save to cache with financial totals and language
        cache = models.InsightsCache(
            user_id=user_id,
            language=language,
            insights=insights,
            financial_snapshot=user_data,
            is_stale=False,
            last_refresh_date=datetime.now(),
            total_income=total_income,
            total_expenses=total_expenses,
            total_loans=total_loans
        )

        db.add(cache)
        db.commit()

        return {
            **insights,
            "metadata": {
                "isCached": False,
                "createdAt": cache.created_at.isoformat(),
                "lastRefreshDate": cache.last_refresh_date.isoformat(),
                "language": language,
                "validityReason": "Generated new insights",
                "financialTotals": {
                    "income": total_income,
                    "expenses": total_expenses,
                    "loans": total_loans
                }
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"[FastAPI] Error in _refresh_insights_internal: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

async def get_user_financial_data(user_id: str, db: Session):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    incomes = db.query(models.Income).filter(models.Income.user_id == user_id).all()
    expenses = db.query(models.Expense).filter(models.Expense.user_id == user_id).all()
    loans = db.query(models.Loan).filter(models.Loan.user_id == user_id).all()
    savings = db.query(models.Saving).filter(models.Saving.user_id == user_id).all()
    savings_goals = db.query(models.SavingsGoal).filter(models.SavingsGoal.user_id == user_id).all()
    financial_freedom = db.query(models.FinancialFreedom).filter(models.FinancialFreedom.userId == user_id).first()
    settings = db.query(models.Settings).filter(models.Settings.user_id == user_id).first()

    # Calculate current year for age-based tax benefits
    current_year = datetime.now().year
    user_age = current_year - settings.birth_year if settings and settings.birth_year else None

    return {
        "settings": {
            "language": settings.language if settings else "en",
            "currency": settings.currency if settings else "PLN",
            # Emergency fund settings
            "emergency_fund_target": settings.emergency_fund_target if settings else 1000,
            "emergency_fund_months": settings.emergency_fund_months if settings else 3,
            # Polish tax settings
            "employment_status": settings.employment_status if settings else None,
            "tax_form": settings.tax_form if settings else None,
            "birth_year": settings.birth_year if settings else None,
            "user_age": user_age,
            "use_authors_costs": settings.use_authors_costs if settings else False,
            # PPK settings
            "ppk_enrolled": settings.ppk_enrolled if settings else None,
            "ppk_employee_rate": settings.ppk_employee_rate if settings else None,
            "ppk_employer_rate": settings.ppk_employer_rate if settings else None,
            "children_count": settings.children_count if settings else 0,
        },
        "incomes": [
            {
                "date": income.date.isoformat(),
                "amount": income.amount,
                "category": income.category,
                "description": income.description,
                "is_recurring": income.is_recurring,
                "employment_type": income.employment_type,
                "gross_amount": income.gross_amount,
                "is_gross": income.is_gross,
            }
            for income in incomes
        ],
        "expenses": [
            {
                "date": expense.date.isoformat(),
                "amount": expense.amount,
                "category": expense.category,
                "description": expense.description,
                "is_recurring": expense.is_recurring
            }
            for expense in expenses
        ],
        "loans": [
            {
                "loan_type": loan.loan_type,
                "description": loan.description,
                "principal_amount": loan.principal_amount,
                "remaining_balance": loan.remaining_balance,
                "interest_rate": loan.interest_rate,
                "monthly_payment": loan.monthly_payment,
                "term_months": loan.term_months,
                "start_date": loan.start_date.isoformat() if loan.start_date else None,
                "overpayment_fee_percent": loan.overpayment_fee_percent,
                "overpayment_fee_waived_until": loan.overpayment_fee_waived_until.isoformat() if loan.overpayment_fee_waived_until else None,
            }
            for loan in loans
        ],
        "savings": [
            {
                "category": saving.category,
                "description": saving.description,
                "amount": saving.amount,
                "date": saving.date.isoformat(),
                "is_recurring": saving.is_recurring,
                "saving_type": saving.saving_type,
                "account_type": saving.account_type,
                "annual_return_rate": saving.annual_return_rate,
                "goal_id": saving.goal_id,
            }
            for saving in savings
        ],
        "savings_goals": [
            {
                "name": goal.name,
                "category": goal.category,
                "target_amount": goal.target_amount,
                "current_amount": goal.current_amount,
                "deadline": goal.deadline.isoformat() if goal.deadline else None,
                "status": goal.status,
                "priority": goal.priority,
            }
            for goal in savings_goals
        ],
        "financial_freedom": {
            "steps": financial_freedom.steps if financial_freedom else None,
            "start_date": financial_freedom.startDate.isoformat() if financial_freedom and financial_freedom.startDate else None,
        } if financial_freedom else None,
    }

async def generate_insights(user_data: dict, api_key: str):
    """
    Generate financial insights using the OpenAI API.
    Aligned with FIRE (Financial Independence, Retire Early) philosophy
    and Dave Ramsey's Baby Steps methodology.

    Args:
        user_data: Dictionary containing user's complete financial data
        api_key: OpenAI API key from environment

    Returns:
        InsightsResponse object with AI-generated insights
    """
    try:
        # Extract all data
        settings = user_data.get("settings", {})
        incomes = user_data.get("incomes", [])
        expenses = user_data.get("expenses", [])
        loans = user_data.get("loans", [])
        savings = user_data.get("savings", [])
        savings_goals = user_data.get("savings_goals", [])
        financial_freedom = user_data.get("financial_freedom", {})

        currency = settings.get("currency", "PLN")
        language = settings.get("language", "en")

        # Calculate basic financial metrics
        total_income = sum(income["amount"] for income in incomes)
        total_expenses = sum(expense["amount"] for expense in expenses)
        total_loan_payments = sum(loan["monthly_payment"] for loan in loans)
        total_loan_balance = sum(loan["remaining_balance"] for loan in loans)

        # Calculate recurring vs one-time breakdown
        recurring_income = sum(i["amount"] for i in incomes if i.get("is_recurring"))
        one_time_income = sum(i["amount"] for i in incomes if not i.get("is_recurring"))
        recurring_expenses = sum(e["amount"] for e in expenses if e.get("is_recurring"))
        one_time_expenses = sum(e["amount"] for e in expenses if not e.get("is_recurring"))

        # One-time expense details for sinking fund analysis
        one_time_expense_details = [
            {"category": e["category"], "description": e.get("description", ""), "amount": e["amount"]}
            for e in expenses if not e.get("is_recurring")
        ]

        # Real monthly surplus based on recurring only
        recurring_monthly_surplus = recurring_income - recurring_expenses - total_loan_payments
        real_savings_rate = (recurring_monthly_surplus / recurring_income * 100) if recurring_income > 0 else 0

        # Calculate monthly balance
        monthly_balance = total_income - total_expenses - total_loan_payments

        # Calculate savings rate (total, including one-time)
        savings_rate = (monthly_balance / total_income * 100) if total_income > 0 else 0

        # Calculate debt-to-income ratio
        debt_to_income = (total_loan_payments / total_income * 100) if total_income > 0 else 0

        # Group expenses by category
        expense_categories = {}
        for expense in expenses:
            category = expense["category"]
            expense_categories[category] = expense_categories.get(category, 0) + expense["amount"]

        # Group incomes by category and employment type
        income_categories = {}
        income_by_employment = {}
        for income in incomes:
            category = income["category"]
            income_categories[category] = income_categories.get(category, 0) + income["amount"]
            emp_type = income.get("employment_type") or "unknown"
            income_by_employment[emp_type] = income_by_employment.get(emp_type, 0) + income["amount"]

        # Calculate savings totals by category and account type (net: deposits - withdrawals)
        savings_by_category = {}
        savings_by_account_type = {}
        total_savings_deposits = 0
        for saving in savings:
            cat = saving["category"]
            acc_type = saving.get("account_type") or "standard"
            amount = saving["amount"]
            if saving.get("saving_type") == "deposit":
                total_savings_deposits += amount
                savings_by_category[cat] = savings_by_category.get(cat, 0) + amount
                savings_by_account_type[acc_type] = savings_by_account_type.get(acc_type, 0) + amount
            elif saving.get("saving_type") == "withdrawal":
                savings_by_category[cat] = savings_by_category.get(cat, 0) - amount
                savings_by_account_type[acc_type] = savings_by_account_type.get(acc_type, 0) - amount

        # Calculate emergency fund status
        # Baby Step 1: Starter emergency fund (fixed target, typically 1000 PLN/USD)
        baby_step_1_target = settings.get("emergency_fund_target", 1000)
        # Baby Step 3: Full emergency fund (3-6 months of expenses)
        emergency_fund_months = settings.get("emergency_fund_months", 3)
        # Liquid savings = emergency_fund + six_month_fund + general (excludes retirement, investment, real_estate)
        liquid_categories = ["emergency_fund", "six_month_fund", "general"]
        emergency_fund_current = sum(savings_by_category.get(cat, 0) for cat in liquid_categories)
        # Use RECURRING expenses for emergency fund target (not total which includes one-time)
        baby_step_3_target = recurring_expenses * emergency_fund_months if recurring_expenses > 0 else baby_step_1_target * emergency_fund_months

        # Pre-calculate differences for BOTH Baby Steps to avoid AI math errors
        baby_step_1_difference = emergency_fund_current - baby_step_1_target
        baby_step_1_complete = emergency_fund_current >= baby_step_1_target
        baby_step_1_status = "complete" if baby_step_1_complete else f"needs {abs(baby_step_1_difference):,.0f} more"

        baby_step_3_difference = emergency_fund_current - baby_step_3_target
        baby_step_3_complete = emergency_fund_current >= baby_step_3_target
        baby_step_3_status = "complete" if baby_step_3_complete else f"needs {abs(baby_step_3_difference):,.0f} more"

        emergency_fund_surplus = baby_step_3_difference if baby_step_3_complete else 0

        # Determine current Baby Step
        baby_steps = financial_freedom.get("steps", []) if financial_freedom else []
        current_baby_step = 0
        baby_steps_summary = []
        for step in baby_steps:
            step_num = step.get("id", 0)
            is_completed = step.get("isCompleted", False)
            progress = step.get("progress", 0)
            baby_steps_summary.append({
                "step": step_num,
                "completed": is_completed,
                "progress": progress,
                "target": step.get("targetAmount"),
                "current": step.get("currentAmount"),
            })
            if not is_completed and current_baby_step == 0:
                current_baby_step = step_num

        # Sort loans by balance for debt snowball analysis (smallest first)
        # NOTE: Only snowball order - avalanche data intentionally excluded per Dave Ramsey methodology
        loans_sorted_by_balance = sorted(loans, key=lambda x: x.get("remaining_balance", 0))

        # Polish tax context
        user_age = settings.get("user_age")
        employment_status = settings.get("employment_status")
        tax_form = settings.get("tax_form")
        ppk_enrolled = settings.get("ppk_enrolled")
        use_authors_costs = settings.get("use_authors_costs")
        children_count = settings.get("children_count", 0)

        # IKE/IKZE limits for 2026 (Poland)
        ike_limit_2026 = 28260  # PLN
        ikze_limit_2026 = 11304  # PLN (16956 for self-employed)
        ikze_limit_b2b = 16956  # PLN

        # Calculate IKE/IKZE year-to-date contributions
        current_year = datetime.now().year
        ike_ytd = sum(
            s["amount"] for s in savings
            if s.get("saving_type") == "deposit"
            and s.get("account_type") == "ike"
            and s.get("date", "")[:4] == str(current_year)
        )
        ikze_ytd = sum(
            s["amount"] for s in savings
            if s.get("saving_type") == "deposit"
            and s.get("account_type") == "ikze"
            and s.get("date", "")[:4] == str(current_year)
        )
        is_self_employed = employment_status in ("b2b", "self_employed", "self-employed")
        ikze_applicable_limit = ikze_limit_b2b if is_self_employed else ikze_limit_2026
        ike_remaining = max(0, ike_limit_2026 - ike_ytd)
        ikze_remaining = max(0, ikze_applicable_limit - ikze_ytd)

        # Calculate FIRE number (25x annual RECURRING expenses = 4% withdrawal rate)
        annual_recurring_expenses = recurring_expenses * 12
        annual_total_expenses = total_expenses * 12
        fire_number = annual_recurring_expenses * 25 if annual_recurring_expenses > 0 else 0
        fire_number_total = annual_total_expenses * 25 if annual_total_expenses > 0 else 0

        # Map language codes
        language_names = {"en": "English", "pl": "Polish", "es": "Spanish"}
        language_name = language_names.get(language, "English")

        # Categorize loans for proper Baby Steps handling
        mortgage_loans = [l for l in loans if l.get("loan_type") == "mortgage"]
        leasing_loans = [l for l in loans if l.get("loan_type") == "leasing"]
        baby_step_2_debts = sorted(
            [l for l in loans if l.get("loan_type") not in ("mortgage", "leasing")],
            key=lambda x: x.get("remaining_balance", 0)
        )
        high_interest_loans = [l for l in loans if l.get("interest_rate", 0) >= 5 and l.get("loan_type") not in ("mortgage", "leasing")]

        # Calculate debt totals by category
        baby_step_2_debt_total = sum(l.get("remaining_balance", 0) for l in baby_step_2_debts)
        baby_step_2_payments = sum(l.get("monthly_payment", 0) for l in baby_step_2_debts)
        mortgage_debt_total = sum(l.get("remaining_balance", 0) for l in mortgage_loans)
        mortgage_payments = sum(l.get("monthly_payment", 0) for l in mortgage_loans)
        leasing_debt_total = sum(l.get("remaining_balance", 0) for l in leasing_loans)
        leasing_payments = sum(l.get("monthly_payment", 0) for l in leasing_loans)

        # Calculate two DTI ratios: Baby Step 2 only vs total
        baby_step_2_dti = (baby_step_2_payments / total_income * 100) if total_income > 0 else 0
        total_dti = (total_loan_payments / total_income * 100) if total_income > 0 else 0

        # Pre-calculate interest savings for high-interest debts (AC4)
        high_interest_savings = []
        for loan in high_interest_loans:
            balance = loan.get("remaining_balance", 0)
            rate = loan.get("interest_rate", 0)
            monthly_payment = loan.get("monthly_payment", 0)
            term_months = loan.get("term_months", 0)

            if balance > 0 and rate > 0 and monthly_payment > 0:
                # Calculate remaining interest at current payment
                monthly_rate = rate / 100 / 12
                remaining_balance = balance
                total_interest_normal = 0
                for _ in range(term_months if term_months > 0 else 360):
                    if remaining_balance <= 0:
                        break
                    interest_payment = remaining_balance * monthly_rate
                    total_interest_normal += interest_payment
                    principal_payment = monthly_payment - interest_payment
                    if principal_payment <= 0:
                        break
                    remaining_balance -= principal_payment

                # Calculate with extra 200 and 500 overpayment
                overpayment_scenarios = []
                for extra in [200, 500]:
                    remaining_balance = balance
                    total_interest_extra = 0
                    months_extra = 0
                    for _ in range(term_months if term_months > 0 else 360):
                        if remaining_balance <= 0:
                            break
                        interest_payment = remaining_balance * monthly_rate
                        total_interest_extra += interest_payment
                        principal_payment = (monthly_payment + extra) - interest_payment
                        if principal_payment <= 0:
                            break
                        remaining_balance -= principal_payment
                        months_extra += 1
                    savings_amount = total_interest_normal - total_interest_extra
                    overpayment_scenarios.append({
                        "extra_monthly": extra,
                        "interest_saved": round(savings_amount, 0),
                        "months_to_payoff": months_extra
                    })

                high_interest_savings.append({
                    "description": loan.get("description", ""),
                    "loan_type": loan.get("loan_type", ""),
                    "balance": balance,
                    "rate": rate,
                    "total_interest_remaining": round(total_interest_normal, 0),
                    "overpayment_scenarios": overpayment_scenarios
                })

        # Build missing data links (AC5 - strict mapping of which links are required)
        missing_data_links = []
        required_links_by_category = {"tax_optimization": [], "savings": [], "baby_steps": [], "fire": []}
        if not user_age:
            missing_data_links.append("age")
            required_links_by_category["tax_optimization"].append("[Uzupełnij profil podatkowy](/settings)")
        if ppk_enrolled is None:
            missing_data_links.append("ppk_status")
            required_links_by_category["tax_optimization"].append("[Uzupełnij profil podatkowy](/settings)")
        if not employment_status:
            missing_data_links.append("employment_status")
            required_links_by_category["tax_optimization"].append("[Uzupełnij profil podatkowy](/settings)")
        if not baby_steps_summary:
            required_links_by_category["baby_steps"].append("[Skonfiguruj Baby Steps](/financial-freedom)")
        # Deduplicate links per category
        for cat in required_links_by_category:
            required_links_by_category[cat] = list(dict.fromkeys(required_links_by_category[cat]))

        # Determine if user has variable/self-employed income (edge case 4)
        has_variable_income = is_self_employed or employment_status in ("freelance", "contract")
        recommended_emergency_months = "6-12" if has_variable_income else f"{emergency_fund_months}"

        # Build comprehensive prompt
        prompt = f"""You are a FIRE (Financial Independence, Retire Early) coach and financial advisor specializing in Dave Ramsey's Baby Steps methodology, with expertise in Polish financial regulations.

PHILOSOPHY:
- Focus on building wealth through intentional living, not deprivation
- Follow the Baby Steps: (1) Starter emergency fund, (2) Pay off all debt (EXCEPT mortgage) using debt snowball, (3) 3-6 months emergency fund, (4) Invest 15% for retirement, (5) College savings, (6) Pay off home early, (7) Build wealth and give
- Use DEBT SNOWBALL method ONLY (smallest balance first) - this is NON-NEGOTIABLE. Do NOT mention or recommend debt avalanche method.
- Savings rate is king - the higher the better for reaching FIRE

CRITICAL RULES - YOU MUST FOLLOW THESE:
1. Do NOT perform any mathematical calculations. ALL numbers are pre-calculated below. Use them exactly as provided.
2. Do NOT recommend debt avalanche. Only debt snowball (smallest balance first).
3. Do NOT suggest paying off leasing early - it's a fixed contract that cannot be prepaid.
4. ALL monetary values, percentages, and differences are pre-calculated. Just reference them in your narrative.

DEBT RULES:
1. **Baby Step 2 debts** (pay off now, smallest to largest): Consumer loans, cash loans, car loans, credit cards, personal loans, installment loans
2. **NEVER include in Baby Step 2**: Mortgages (Baby Step 6), Leasing (fixed contracts that CANNOT be prepaid early)
3. **Leasing**: These are FIXED CONTRACTS. The user CANNOT prepay them early. Never suggest paying off leasing early. Explain: "fixed contract, cannot be prepaid early."
4. **Mortgage**: Only addressed in Baby Step 6 AFTER completing Baby Steps 1-5. If user is on Baby Step 2-5, do NOT discuss mortgage payoff.
5. **Overpayment fees**: In Poland, since 2022 banks cannot charge prepayment fees for first 3 years. Check overpayment_fee_percent and overpayment_fee_waived_until fields.

CURRENT BABY STEP: {current_baby_step if current_baby_step > 0 else "Not started or data unavailable"}

BABY STEPS PROGRESS:
{json.dumps(baby_steps_summary, indent=2) if baby_steps_summary else "No Baby Steps data available"}

USER CONTEXT (Polish market):
- Age: {user_age if user_age else "UNKNOWN"}
- Employment: {employment_status or "UNKNOWN"} ({tax_form or "Unknown"} tax form)
- PPK enrolled: {ppk_enrolled if ppk_enrolled is not None else "UNKNOWN"}
- Uses author's costs (50% KUP, UoP/Dzieło only): {use_authors_costs}{"" if not is_self_employed else " (NOTE: 50% KUP does NOT apply to B2B - user should deduct actual business expenses via invoices)"}
- Children: {children_count}
- Currency: {currency}
- Has variable/self-employed income: {"YES - recommend higher emergency fund (6-12 months)" if has_variable_income else "NO"}

MISSING DATA - MANDATORY LINKS (you MUST include these links in actionItems of the specified categories):
{json.dumps(required_links_by_category, indent=2)}
When any of these links apply, add them to the actionItems array of insights in that category. Do NOT put links inside description text.

POLISH TAX OPTIMIZATION OPPORTUNITIES:
- Youth tax relief (ulga dla młodych): Available if under 26 years old, income up to 85,528 PLN/year tax-free
{"- ⚠️ User is " + str(user_age) + " years old - youth tax relief EXPIRES at 26! Prominently feature this." if user_age and user_age <= 26 else ""}
- IKE (Individual Retirement Account): 2026 limit = {ike_limit_2026} PLN, tax-free capital gains
- IKZE (Individual Retirement Security Account): 2026 limit = {ikze_applicable_limit} PLN {"(B2B higher limit)" if is_self_employed else "(standard limit)"}, tax-deductible contributions
- PPK: Employer matches contributions, free money if enrolled
{"- Author's costs (KUP 50%): 50% of income tax-deductible for creative work (UoP/Umowa o dzieło ONLY)" if not is_self_employed else ""}

TAX OPTIMIZATION RULES BY EMPLOYMENT TYPE:
- B2B (linear tax / ryczałt): NO "50% KUP" — that applies ONLY to UoP/Umowa o dzieło. On B2B, deduct actual business expenses from invoices (koszty uzyskania przychodu z faktur). Suggest checking if personal car expenses, equipment, home office, phone, internet could be legitimate business costs (potential ~23% VAT + income tax recovery).
- B2B health insurance: 9% of average salary base, partially deductible from tax (4.9% for linear tax)
- UoP (employment contract): Standard KUP 250 PLN/month or 300 PLN if commuting. 50% KUP available for creative/IP work.
- If user has children: Joint filing (rozliczenie wspólne) NOT possible on linear tax (podatek liniowy). If beneficial, mention considering switching to skala podatkowa next year.
- Tax leak detection: If user has B2B AND personal transportation/equipment/subscriptions expenses, ask whether these are already classified as business costs.

SINKING FUNDS DETECTION:
When you see large one-time expenses (vacations, car service/repairs, insurance, medical, gifts), calculate their estimated annual cost and suggest a monthly "sinking fund" amount. Example: "You spent X PLN on one-time expenses this month. If these recur annually, setting aside Y PLN/month to a dedicated sub-account would smooth out cash flow and avoid surprises."
This helps the user plan for predictable irregular expenses instead of being surprised by them each time.

FINANCIAL SUMMARY (PRE-CALCULATED - use these exact numbers):
- Monthly Income (net): {total_income:,.2f} {currency}
- Monthly Expenses: {total_expenses:,.2f} {currency}
- Monthly Loan Payments (all): {total_loan_payments:,.2f} {currency}
- Monthly Balance (income - expenses - loan payments): {monthly_balance:,.2f} {currency}
- Cash Savings Rate (this month, incl. one-time): {savings_rate:.1f}%
- Note: Debt principal payments of ~{total_loan_payments:,.0f} {currency}/month also build net worth by reducing debt, but are not counted in savings rate.
- Baby Step 2 DTI (consumer debt only): {baby_step_2_dti:.1f}% (only non-mortgage, non-leasing debt payments / income)
- Total DTI (all debt): {total_dti:.1f}% (all loan payments / income)
- Total All Debt: {total_loan_balance:,.2f} {currency}
- Baby Step 2 Debt (to pay off now): {baby_step_2_debt_total:,.2f} {currency}
- Mortgage Debt (Baby Step 6): {mortgage_debt_total:,.2f} {currency}
- Leasing Debt (fixed contracts, CANNOT prepay): {leasing_debt_total:,.2f} {currency}
- FIRE Number (25x annual RECURRING expenses): {fire_number:,.0f} {currency}
- FIRE Number reference (25x ALL expenses incl. one-time): {fire_number_total:,.0f} {currency}

REAL MONTHLY BUDGET (excluding one-time expenses):
- Recurring Income: {recurring_income:,.2f} {currency}
- Recurring Expenses: {recurring_expenses:,.2f} {currency}
- One-Time Income this month: {one_time_income:,.2f} {currency}
- One-Time Expenses this month: {one_time_expenses:,.2f} {currency}
- Real Monthly Surplus (recurring income - recurring expenses - loan payments): {recurring_monthly_surplus:,.2f} {currency}
- Real Savings Rate (based on recurring only): {real_savings_rate:.1f}%
IMPORTANT: Use RECURRING expenses for emergency fund calculation, Baby Steps progress, and FIRE number.
Use TOTAL expenses only for current month snapshot. When savings rate differs significantly between total and recurring,
explain that one-time expenses distort the current month picture.

ONE-TIME EXPENSES THIS MONTH (for sinking fund analysis):
{json.dumps(one_time_expense_details, indent=2) if one_time_expense_details else "None"}

EMERGENCY FUND STATUS (PRE-CALCULATED - use these EXACT values, do NOT recalculate):
- Liquid savings (emergency_fund + six_month_fund + general): {emergency_fund_current:,.0f} {currency}
- BABY STEP 1 (starter emergency fund):
  Target: {baby_step_1_target:,.0f} {currency}
  Current liquid savings: {emergency_fund_current:,.0f} {currency}
  Status: {baby_step_1_status}
  Complete: {"YES ✓" if baby_step_1_complete else "NO - needs " + f"{abs(baby_step_1_difference):,.0f}" + " " + currency + " more"}

- BABY STEP 3 (full emergency fund = {emergency_fund_months} months of RECURRING expenses):
  Based on: {recurring_expenses:,.0f} {currency}/month RECURRING expenses (NOT {total_expenses:,.0f} total)
  Target: {baby_step_3_target:,.0f} {currency}
  Current liquid savings: {emergency_fund_current:,.0f} {currency}
  Status: {baby_step_3_status}
  Complete: {"YES ✓" if baby_step_3_complete else "NO - needs " + f"{abs(baby_step_3_difference):,.0f}" + " " + currency + " more"}
{"  Surplus: " + f"{emergency_fund_surplus:,.0f}" + " " + currency + " above target. Consider redirecting excess to Baby Step 4 (investing) or Baby Step 6 (mortgage)." if emergency_fund_surplus > 0 else ""}
{" Recommended months for this user: " + recommended_emergency_months + " months (variable income detected)" if has_variable_income else ""}

INCOME BY EMPLOYMENT TYPE:
{json.dumps(income_by_employment, indent=2)}

EXPENSE BREAKDOWN:
{json.dumps(expense_categories, indent=2)}

BABY STEP 2 DEBTS (PRE-SORTED smallest to largest - this IS the Debt Snowball order, recommend paying in THIS exact order):
{json.dumps(baby_step_2_debts, indent=2) if baby_step_2_debts else "No Baby Step 2 debts - user is debt-free (excluding mortgage/leasing)!"}

{"ONLY MORTGAGE REMAINING - User has completed Baby Steps 1-5 debt elimination! Focus entirely on Baby Step 6 mortgage payoff strategy." if not baby_step_2_debts and mortgage_loans and baby_step_1_complete else ""}

MORTGAGE (Baby Step 6 - {"ACTIVE: user has completed earlier steps" if not baby_step_2_debts and baby_step_1_complete else "IGNORE until Baby Steps 1-5 completed"}):
{json.dumps(mortgage_loans, indent=2) if mortgage_loans else "No mortgage"}

LEASING (FIXED CONTRACTS - CANNOT be prepaid early, do NOT include in any payoff strategy):
{json.dumps(leasing_loans, indent=2) if leasing_loans else "No leasing"}

HIGH INTEREST DEBTS (>5%) WITH PRE-CALCULATED OVERPAYMENT SAVINGS:
{json.dumps(high_interest_savings, indent=2) if high_interest_savings else "No high-interest Baby Step 2 debts"}
Use the overpayment_scenarios above to recommend specific extra monthly payments. Reference the exact interest_saved amounts.

SAVINGS BY CATEGORY:
{json.dumps(savings_by_category, indent=2)}

SAVINGS BY ACCOUNT TYPE (IKE/IKZE/PPK/Standard):
{json.dumps(savings_by_account_type, indent=2)}

IKE/IKZE CONTRIBUTION STATUS {current_year} (PRE-CALCULATED):
- IKE: Contributed {ike_ytd:,.0f} / {ike_limit_2026:,.0f} PLN this year. Remaining room: {ike_remaining:,.0f} PLN
- IKZE: Contributed {ikze_ytd:,.0f} / {ikze_applicable_limit:,.0f} PLN this year. Remaining room: {ikze_remaining:,.0f} PLN
- Recommendation order: First max IKZE (tax deduction benefit), then IKE (if budget allows)

SAVINGS GOALS:
{json.dumps(savings_goals, indent=2)}

Generate insights in these 5 categories based on the user's current Baby Step:
1. **baby_steps** - Where they are in the Baby Steps journey, what to focus on NOW
2. **debt** - Debt snowball payoff strategy, specific next debt to target
3. **savings** - Emergency fund progress, IKE/IKZE contribution room and optimization
4. **fire** - FIRE number, savings rate analysis, time to financial independence
5. **tax_optimization** - Polish-specific tax opportunities based on their situation

HERO DASHBOARD (generate this object alongside categories):
Generate a "hero_dashboard" object with:
- greeting: One warm sentence summarizing the user's financial health (max 20 words, in response language)
- health_status: "excellent" if real surplus >30% of income, "good" if >10%, "warning" if >0%, "critical" if negative
- monthly_cost_of_living: RECURRING expenses only (use {recurring_expenses:,.2f} {currency})
- monthly_income: recurring income (use {recurring_income:,.2f} {currency})
- monthly_surplus: recurring income - recurring expenses - loan payments (use {recurring_monthly_surplus:,.2f} {currency})
- fire_progress_percent: (current savings / FIRE number) * 100, rounded to 1 decimal
- fire_target: FIRE number (use {fire_number:,.0f} {currency})
- budget_distortion: If one-time expenses > 30% of total expenses this month, set is_distorted=true. Include one_time_total, explanation (why this month looks different), and corrected_surplus (what surplus would be without one-time expenses)
- top3_moves: The 3 most impactful financial actions, specific to this user's data, with concrete {currency} amounts. Each must have: title (short), description (1-2 sentences), impact (specific savings/gain like "Oszczędzisz X PLN rocznie"), icon_type (one of: mortgage, savings, investment, budget, tax, emergency)

TOP 3 MOVES RULES:
- Each move MUST have a specific {currency} impact calculated from the user's data (not generic advice)
- Prioritize by impact: highest {currency} savings/gain first
- Good: "Nadpłać hipotekę 3000 zł/mies → oszczędność 22k odsetek" — Bad: "Śledź wydatki"
- Never suggest generic things like "track expenses" or "read a book"

CRITICAL GUIDELINES:
- Be specific and actionable - tell them EXACTLY what to do next
- Reference the pre-calculated numbers exactly as provided - do NOT recalculate anything
- Use DEBT SNOWBALL only (smallest balance first) - NEVER mention debt avalanche
- If they're on Baby Step 2, focus ONLY on Baby Step 2 debts (exclude mortgage and leasing)
- NEVER suggest paying off leasing early - explain it's a fixed contract
- For mortgage - ONLY discuss in Baby Step 6 context after earlier steps are done
- For high-interest debts: use the PRE-CALCULATED overpayment savings scenarios and recommend specific monthly overpayment amounts
- If user has high-interest debt AND savings beyond emergency fund, suggest using excess savings for debt payoff
- EMERGENCY FUND: Use the EXACT pre-calculated values. Baby Step 1 and Baby Step 3 are SEPARATE targets.
- SAVINGS RATE: Present as "Cash savings rate" and note that debt principal payments also build net worth but are not included
- DTI: Reference baby_step_2_dti for debt discussion, total_dti for overall health assessment
- LINKS: Include mandatory links from MISSING DATA section in the actionItems of the specified categories
- IKE/IKZE: Reference exact remaining contribution room for the current year
- {"User has NO debts but negative cash flow - focus on expense reduction and Baby Step 1 emergency fund urgency" if not loans and monthly_balance < 0 else ""}
- {"User has emergency fund surplus of " + f"{emergency_fund_surplus:,.0f}" + " " + currency + " - suggest redirecting to investing or mortgage, do NOT recommend reducing emergency fund" if emergency_fund_surplus > 0 else ""}
- If they're past Baby Step 3, focus on IKE/IKZE optimization and FIRE progress
- Celebrate achievements but be direct about what needs improvement
- For Polish users, always mention relevant tax benefits they might be missing

RESPOND IN: {language_name}
{"Use Polish number format: space as thousands separator, comma as decimal (e.g., 12 345,67 zł). Use Polish date format (DD.MM.YYYY)." if language == "pl" else ""}

JSON Response Structure (respond with ONLY valid JSON, no other text):
{{
  "categories": {{
    "baby_steps": [
      {{
        "type": "observation|recommendation|alert|achievement",
        "title": "Insight title",
        "description": "Detailed description with specific numbers from the data above. Do NOT include links here.",
        "priority": "high|medium|low",
        "actionItems": ["Specific action 1", "Markdown links go HERE in actionItems, e.g. [Link text](/path)"],
        "metrics": [
          {{"label": "Metric", "value": "Value with currency", "trend": "up|down|stable"}}
        ]
      }}
    ],
    "debt": [...],
    "savings": [...],
    "fire": [...],
    "tax_optimization": [...]
  }},
  "status": {{
    "baby_steps": "good|ok|can_be_improved|bad",
    "debt": "good|ok|can_be_improved|bad",
    "savings": "good|ok|can_be_improved|bad",
    "fire": "good|ok|can_be_improved|bad",
    "tax_optimization": "good|ok|can_be_improved|bad"
  }},
  "currentBabyStep": {current_baby_step},
  "fireNumber": {fire_number},
  "savingsRate": {savings_rate:.1f},
  "realSavingsRate": {real_savings_rate:.1f},
  "hero_dashboard": {{
    "greeting": "One warm sentence about financial health",
    "health_status": "excellent|good|warning|critical",
    "monthly_cost_of_living": 0,
    "monthly_income": 0,
    "monthly_surplus": 0,
    "fire_progress_percent": 0.0,
    "fire_target": 0,
    "budget_distortion": {{
      "is_distorted": false,
      "one_time_total": 0,
      "explanation": "Why this month looks different (or empty if not distorted)",
      "corrected_surplus": 0
    }},
    "top3_moves": [
      {{
        "title": "Short action title",
        "description": "1-2 sentence description",
        "impact": "Specific PLN impact",
        "icon_type": "mortgage|savings|investment|budget|tax|emergency"
      }}
    ]
  }}
}}"""

        # Make the API call to OpenAI API with retry for transient errors
        system_prompt = "You are a FIRE (Financial Independence, Retire Early) coach specializing in Dave Ramsey's Baby Steps methodology and Polish financial regulations. You provide direct, actionable insights focused on debt elimination, emergency fund building, and wealth accumulation. CRITICAL: 1) Respond with valid JSON ONLY - no markdown code blocks, no explanatory text. 2) Do NOT perform any mathematical calculations - all numbers are pre-calculated in the prompt, use them exactly. 3) Never recommend debt avalanche - only debt snowball. 4) Place links in actionItems arrays, never in description text. 5) ALWAYS distinguish between recurring and one-time expenses in your analysis. Emergency fund targets and FIRE numbers are based on RECURRING expenses only. 6) For B2B users, NEVER suggest '50% KUP' - that's for UoP/Dzieło only. B2B deducts actual business costs from invoices."

        max_retries = 2
        response = None
        for attempt in range(max_retries + 1):
            async with httpx.AsyncClient(timeout=90.0) as client:
                response = await client.post(
                    "https://api.openai.com/v1/chat/completions",
                    headers={
                        "Authorization": f"Bearer {api_key}",
                        "Content-Type": "application/json",
                    },
                    json={
                        "model": "gpt-4.1-mini",
                        "max_tokens": 5000,
                        "response_format": {"type": "json_object"},
                        "messages": [
                            {"role": "system", "content": system_prompt},
                            {"role": "user", "content": prompt}
                        ],
                    },
                )

            # Retry on 500/502/503 (transient) with backoff
            if response.status_code in (500, 502, 503) and attempt < max_retries:
                import asyncio
                wait_time = 5 * (attempt + 1)  # 5s, 10s
                print(f"[OpenAI API] {response.status_code} error, retrying in {wait_time}s (attempt {attempt + 1}/{max_retries})")
                await asyncio.sleep(wait_time)
                continue
            break

        # Check if the request was successful
        if response.status_code != 200:
            print(f"[OpenAI API] Error: {response.status_code} - {response.text}")

            if response.status_code == 401:
                raise Exception("OpenAI API error: 401 - Invalid API key.")
            elif response.status_code == 429:
                raise Exception("OpenAI API error: 429 - Rate limit exceeded. Please try again later.")
            else:
                raise Exception(f"OpenAI API error: {response.status_code}")

        # Parse the response
        openai_response = response.json()
        choices = openai_response.get("choices", [])

        if not choices:
            raise Exception("Empty response from OpenAI API")

        # Extract text from the first choice
        text_content = choices[0].get("message", {}).get("content", "")

        text_content = text_content.strip()

        # Try to parse the JSON response
        try:
            # Find JSON in the response (it might be wrapped in markdown code blocks)
            json_match = re.search(r'```json\s*([\s\S]*?)\s*```|({[\s\S]*})', text_content)
            if json_match:
                json_str = json_match.group(1) or json_match.group(2)
                insights_data = json.loads(json_str)
            else:
                insights_data = json.loads(text_content)

            # Validate required fields (AC9)
            required_categories = ["baby_steps", "debt", "savings", "fire", "tax_optimization"]
            categories = insights_data.get("categories", {})
            for cat in required_categories:
                if cat not in categories:
                    categories[cat] = []
            insights_data["categories"] = categories

            status = insights_data.get("status", {})
            for cat in required_categories:
                if cat not in status:
                    status[cat] = "ok"
            insights_data["status"] = status

            # Ensure top-level metrics are present
            if "currentBabyStep" not in insights_data:
                insights_data["currentBabyStep"] = current_baby_step
            if "fireNumber" not in insights_data:
                insights_data["fireNumber"] = fire_number
            if "savingsRate" not in insights_data:
                insights_data["savingsRate"] = round(savings_rate, 1)
            if "realSavingsRate" not in insights_data:
                insights_data["realSavingsRate"] = round(real_savings_rate, 1)

            # Add metadata
            insights_data["metadata"] = {
                "generatedAt": datetime.now().isoformat(),
                "source": "gpt-4.1-mini"
            }

            return insights_data

        except json.JSONDecodeError as e:
            print(f"[OpenAI API] JSON parse error: {e}")
            print(f"[OpenAI API] Response content: {text_content[:500]}")
            raise Exception(f"Failed to parse OpenAI API response: {e}")

    except Exception as e:
        print(f"[OpenAI API] Error generating insights: {str(e)}")

        # Return a fallback response in case of error
        sample_insight = {
            "type": "observation",
            "title": "API Error",
            "description": f"We encountered an error while generating insights: {str(e)}. Please try again later.",
            "priority": "medium",
            "actionItems": ["Try again later", "Contact support if the issue persists"],
            "metrics": [
                {
                    "label": "Error",
                    "value": "API call failed",
                    "trend": "stable"
                }
            ]
        }

        return {
            "categories": {
                "baby_steps": [sample_insight],
                "debt": [sample_insight],
                "savings": [sample_insight],
                "fire": [sample_insight],
                "tax_optimization": [sample_insight]
            },
            "status": {
                "baby_steps": "ok",
                "debt": "ok",
                "savings": "ok",
                "fire": "ok",
                "tax_optimization": "ok"
            },
            "currentBabyStep": 0,
            "fireNumber": 0,
            "savingsRate": 0,
            "metadata": {
                "generatedAt": datetime.now().isoformat(),
                "source": "error_fallback",
                "error": str(e)
            }
        }

class SavingsGoalCreate(BaseModel):
    name: str
    category: str = "general"
    target_amount: float
    deadline: Optional[str] = None  # ISO date string
    priority: int = 0
    status: str = "active"
    icon: Optional[str] = None
    color: Optional[str] = None
    notes: Optional[str] = None

@app.post("/users/{user_id}/savings-goals")
async def create_savings_goal(
    user_id: str,
    goal_data: SavingsGoalCreate,
    current_user: models.User = Depends(get_authenticated_user),
    db: Session = Depends(database.get_db),
):
    validate_user_access(user_id, current_user)
    goal = models.SavingsGoal(
        user_id=current_user.household_id,
        name=goal_data.name,
        category=goal_data.category,
        target_amount=goal_data.target_amount,
        deadline=datetime.fromisoformat(goal_data.deadline.replace("Z", "+00:00")).date() if goal_data.deadline else None,
        priority=goal_data.priority,
        status=goal_data.status,
        icon=goal_data.icon,
        color=goal_data.color,
        notes=goal_data.notes,
    )
    db.add(goal)
    db.commit()
    db.refresh(goal)
    return {
        "id": goal.id,
        "name": goal.name,
        "category": goal.category,
        "target_amount": goal.target_amount,
        "deadline": goal.deadline.isoformat() if goal.deadline else None,
        "priority": goal.priority,
        "status": goal.status,
    }

@app.get("/")
async def root():
    return {"message": "Welcome to the Home Budget API"}
